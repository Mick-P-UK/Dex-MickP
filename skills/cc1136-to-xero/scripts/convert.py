#!/usr/bin/env python3
"""
Halifax Clarity 1136 PDF statements -> Xero import CSV converter
================================================================

Parses a folder of monthly Halifax Clarity 1136 PDF statements, filters
to a given year-end window, builds a running balance from a supplied
opening balance, and writes two output files:

  1. A 11-column CSV ready for Xero bank statement import (matching
     Mick's existing YE 2024 template)
  2. An audit XLSX with three sheets (Transactions, Stmt
     Reconciliation, YE Summary)

Reconciles every parsed statement to Halifax's own reported figures
and every interior statement boundary to Halifax's New balance.
Fails loudly if any reconciliation check breaks.

Usage:
    python3 convert.py \\
        --input-dir <folder-with-PDFs> \\
        --output-dir <where-to-write-outputs> \\
        --opening-balance <prior-YE-closing-balance> \\
        --ye-year <year> \\
        --date-today <YYYY.MM.DD>

Example:
    python3 convert.py \\
        --input-dir "/path/to/halifax-pdfs" \\
        --output-dir "C:/Vaults/DB-Accounts-CW/YE_2025.11.30/workings" \\
        --opening-balance -1786.33 \\
        --ye-year 2025 \\
        --date-today 2026.05.28
"""

from __future__ import annotations

import argparse
import csv
import glob
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber is required. Install with: pip install pdfplumber")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------
# Halifax PDF parsing
# ---------------------------------------------------------------------

MONTHS = {
    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4,
    "MAY": 5, "JUNE": 6, "JULY": 7, "AUGUST": 8,
    "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12,
}
MONTHS_TITLE = {k.title(): v for k, v in MONTHS.items()}

LAYOUT_OLD = dict(
    card    = None,
    txn     = (40, 145),
    posted  = (145, 235),
    desc    = (235, 355),
    loc     = (355, 420),
    state   = (420, 500),
    amount  = (500, 580),
)
LAYOUT_NEW = dict(
    card    = (40, 90),
    txn     = (90, 180),
    posted  = (180, 245),
    desc    = (245, 355),
    loc     = (355, 420),
    state   = (420, 500),
    amount  = (500, 580),
)

STATEMENT_DATE_RE = re.compile(
    r"Your credit card statement\s+(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})"
)
PLAIN_DATE_RE = re.compile(
    r"(\d{1,2})\s+(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})"
)


@dataclass
class Transaction:
    txn_date: date
    posted_date: date
    description: str
    location: str = ""
    state: str = ""
    fx_amount: str = ""
    amount_gbp: Decimal = Decimal("0.00")  # signed


@dataclass
class Statement:
    source_pdf: str
    statement_date: date
    previous_balance: Decimal
    payments_received: Decimal
    new_charges: Decimal
    new_balance: Decimal
    transactions: list = field(default_factory=list)


def _parse_statement_date(page1_text: str) -> date:
    m = STATEMENT_DATE_RE.search(page1_text)
    if m:
        return date(int(m.group(3)), MONTHS_TITLE[m.group(2).title()], int(m.group(1)))
    m2 = PLAIN_DATE_RE.search(page1_text)
    if m2:
        return date(int(m2.group(3)), MONTHS_TITLE[m2.group(2)], int(m2.group(1)))
    raise ValueError("Could not find statement date on page 1")


def _parse_money(s: str) -> Decimal:
    return Decimal(s.replace("GBP", "").replace("£", "").replace(",", "").strip())


def _parse_summary(page1_text: str) -> dict:
    out = {}
    for line in page1_text.splitlines():
        s = line.strip()
        if s.lower().startswith("previous balance"):
            m = re.search(r"([\d,]+\.\d{2})", s)
            if m: out["previous_balance"] = _parse_money(m.group(1))
        elif s.lower().startswith("payments received"):
            m = re.search(r"([\d,]+\.\d{2})", s)
            if m: out["payments_received"] = _parse_money(m.group(1))
        elif s.lower().startswith("new transactions"):
            m = re.search(r"([\d,]+\.\d{2})", s)
            if m: out["new_charges"] = _parse_money(m.group(1))
    m = re.search(r"Your new balance\s*\n?\s*£?\s*([\d,]+\.\d{2})", page1_text)
    if m:
        out["new_balance"] = _parse_money(m.group(1))
    else:
        out["new_balance"] = (
            out.get("previous_balance", Decimal(0))
            - out.get("payments_received", Decimal(0))
            + out.get("new_charges", Decimal(0))
        )
    return out


def _detect_layout(words):
    """Auto-detect OLD vs NEW layout by inspecting page-3 header row."""
    for w in words:
        if 170 <= w["top"] <= 195 and w["text"].lower() == "ending":
            return LAYOUT_NEW
    return LAYOUT_OLD


def _which_column(x0, layout):
    for name, bounds in layout.items():
        if bounds is None:
            continue
        lo, hi = bounds
        if lo <= x0 < hi:
            return name
    return None


def _empty_cells():
    return {"card": [], "txn": [], "posted": [], "desc": [], "loc": [], "state": [], "amount": []}


def _group_rows(words, layout, tol=3.0):
    if not words:
        return []
    ws = sorted(words, key=lambda w: (w["top"], w["x0"]))
    rows = []
    current_top = ws[0]["top"]
    current = {"top": current_top, "cells": _empty_cells()}
    for w in ws:
        if abs(w["top"] - current_top) > tol:
            rows.append(current)
            current_top = w["top"]
            current = {"top": current_top, "cells": _empty_cells()}
        col = _which_column(w["x0"], layout)
        if col is None:
            continue
        current["cells"][col].append(w["text"])
    rows.append(current)
    return rows


def _infer_year(txn_month, stmt_date):
    if txn_month == 12 and stmt_date.month == 1:
        return stmt_date.year - 1
    return stmt_date.year


def _parse_transactions(rows, stmt_date):
    txns = []
    for r in rows:
        cells = r["cells"]
        txn_cell = cells["txn"]
        amt_cell = cells["amount"]
        desc_cell = cells["desc"]

        if not txn_cell:
            if txns and (desc_cell or cells["loc"] or cells["state"]):
                parts = []
                if desc_cell: parts.append(" ".join(desc_cell))
                if cells["loc"]: parts.append(" ".join(cells["loc"]))
                if cells["state"]: parts.append(" ".join(cells["state"]))
                joined = " ".join(parts)
                m = re.match(r"^([\d,]+\.\d{2})\s+([A-Z]{3})\s+@\s+([\d.]+)$", joined)
                if m:
                    txns[-1].fx_amount = f"{m.group(1)} {m.group(2)} @ {m.group(3)}"
            continue

        if len(txn_cell) < 2:
            continue
        try:
            d1 = int(txn_cell[0]); m1 = MONTHS[txn_cell[1].upper()]
        except (ValueError, KeyError):
            continue

        posted_cell = cells["posted"]
        if len(posted_cell) >= 2:
            try:
                d2 = int(posted_cell[0]); m2 = MONTHS[posted_cell[1].upper()]
            except (ValueError, KeyError):
                d2, m2 = d1, m1
        else:
            d2, m2 = d1, m1

        if not amt_cell:
            continue
        amt_text = " ".join(amt_cell)
        am = re.match(r"^([\d,]+\.\d{2})\s*(CR)?$", amt_text)
        if not am:
            continue
        amount = Decimal(am.group(1).replace(",", ""))
        is_credit = bool(am.group(2))

        y_txn = _infer_year(m1, stmt_date)
        y_posted = _infer_year(m2, stmt_date)
        if stmt_date.month == 1 and m1 == 12 and m2 == 1:
            y_txn = stmt_date.year - 1
            y_posted = stmt_date.year

        try:
            txn_date = date(y_txn, m1, d1)
            posted_date = date(y_posted, m2, d2)
        except ValueError:
            continue

        signed = amount if is_credit else -amount
        txns.append(Transaction(
            txn_date=txn_date,
            posted_date=posted_date,
            description=" ".join(desc_cell),
            location=" ".join(cells["loc"]),
            state=" ".join(cells["state"]),
            amount_gbp=signed,
        ))
    return txns


def parse_statement(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        page1_text = pdf.pages[0].extract_text() or ""
        stmt_date = _parse_statement_date(page1_text)
        summary = _parse_summary(page1_text)
        page3 = pdf.pages[2]
        words = page3.extract_words()
        layout = _detect_layout(words)
        rows = _group_rows(words, layout)
        txns = _parse_transactions(rows, stmt_date)
        # Some statements may overflow to page 4 - peek if it looks like
        # the txn table continued.
        if len(pdf.pages) >= 4:
            p4_text = pdf.pages[3].extract_text() or ""
            if any(re.match(r"\d{1,2}\s+(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)", ln.strip()) for ln in p4_text.splitlines()):
                p4_words = pdf.pages[3].extract_words()
                p4_rows = _group_rows(p4_words, layout)
                txns += _parse_transactions(p4_rows, stmt_date)
    return Statement(
        source_pdf=pdf_path,
        statement_date=stmt_date,
        previous_balance=summary.get("previous_balance", Decimal("0")),
        payments_received=summary.get("payments_received", Decimal("0")),
        new_charges=summary.get("new_charges", Decimal("0")),
        new_balance=summary.get("new_balance", Decimal("0")),
        transactions=txns,
    )


# ---------------------------------------------------------------------
# CSV / XLSX builders
# ---------------------------------------------------------------------

def _compose_description(t):
    if t.fx_amount:
        return f"{t.description} ({t.fx_amount})"
    return t.description


def build_rows(statements, opening_balance, ye_open_date, ye_close_date):
    # Annotate each txn with its source statement
    all_items = []
    seen_keys = set()
    for s in sorted(statements, key=lambda s: s.statement_date):
        for t in s.transactions:
            k = (t.txn_date, t.posted_date, t.amount_gbp, t.description, t.location)
            if k in seen_keys:
                continue
            seen_keys.add(k)
            all_items.append((t, s))

    ye_items = [(t, s) for (t, s) in all_items
                if ye_open_date <= t.txn_date <= ye_close_date]
    ye_items.sort(key=lambda ts: (ts[0].posted_date, ts[0].txn_date))

    last_in_stmt = {}
    for i, (t, s) in enumerate(ye_items):
        last_in_stmt[s.statement_date] = i

    rows = []
    running = opening_balance
    for i, (t, s) in enumerate(ye_items):
        amount = t.amount_gbp
        check_bal = ""
        notes = ""
        if i == 0:
            check_bal = str(opening_balance)
            notes = "<- Start of Yr Bal."
        running = running + amount
        if i != len(ye_items) - 1 and last_in_stmt.get(s.statement_date) == i:
            check_bal = str(running)
            notes = "<- New Balance"
        if i == len(ye_items) - 1:
            check_bal = str(running)
            notes = "<- EofYr_Balance"
        rows.append({
            "Date": t.txn_date.strftime("%d/%m/%Y"),
            "Amount": str(amount),
            "DateEntered": t.posted_date.strftime("%d/%m/%Y"),
            "Description": _compose_description(t),
            "Descr2": t.location,
            "Descr3": t.state,
            "Amount1": str(abs(amount)),
            "Amount2": str(amount),
            "Balance": str(running),
            "CheckBal": check_bal,
            "Notes": notes,
            "_stmt_date": s.statement_date.isoformat(),
            "_stmt_newbal": str(s.new_balance),
            "_fx": t.fx_amount,
        })
    return rows


def write_csv(rows, path):
    headers = [
        "*Date", "*Amount", "Date entered", "Description", "Descr.2",
        "Descr.3", "Amount1 GBP", "Amount2", "Balance GBP",
        "Check Bal GBP", "Notes",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([
                r["Date"], r["Amount"], r["DateEntered"], r["Description"],
                r["Descr2"], r["Descr3"], r["Amount1"], r["Amount2"],
                r["Balance"], r["CheckBal"], r["Notes"],
            ])


def write_audit_xlsx(rows, stmts, path, opening_balance, ye_open_date, ye_close_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = [
        "*Date", "*Amount", "Date entered", "Description", "Descr.2",
        "Descr.3", "Amount1 GBP", "Amount2", "Balance GBP",
        "Check Bal GBP", "Notes", "Stmt date", "Stmt New Bal", "FX detail",
    ]
    ws.append(headers)
    bold = Font(bold=True)
    yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    grey = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    for c in ws[1]:
        c.font = bold
        c.fill = grey
    for r in rows:
        ws.append([
            r["Date"], float(r["Amount"]), r["DateEntered"], r["Description"],
            r["Descr2"], r["Descr3"], float(r["Amount1"]), float(r["Amount2"]),
            float(r["Balance"]),
            float(r["CheckBal"]) if r["CheckBal"] else None,
            r["Notes"], r["_stmt_date"], float(r["_stmt_newbal"]), r["_fx"],
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[10].value:
            for c in row:
                c.fill = yellow

    ws2 = wb.create_sheet("Stmt Reconciliation")
    ws2.append([
        "Statement date", "Halifax Prev bal", "Halifax Payments",
        "Halifax New charges", "Halifax New bal", "Txn count (in stmt)",
        "Sum of stmt txns", "Reconciliation diff",
    ])
    for c in ws2[1]:
        c.font = bold
        c.fill = grey
    for s in sorted(stmts, key=lambda s: s.statement_date):
        total = sum((t.amount_gbp for t in s.transactions), Decimal("0"))
        expected = s.payments_received - s.new_charges
        diff = total - expected
        ws2.append([
            s.statement_date.isoformat(),
            float(s.previous_balance),
            float(s.payments_received),
            float(s.new_charges),
            float(s.new_balance),
            len(s.transactions),
            float(total),
            float(diff),
        ])

    ws3 = wb.create_sheet("YE Summary")
    ws3.append(["Item", "Value"])
    for c in ws3[1]:
        c.font = bold
        c.fill = grey
    ye_debits = sum((Decimal(r["Amount"]) for r in rows if Decimal(r["Amount"]) < 0), Decimal("0"))
    ye_credits = sum((Decimal(r["Amount"]) for r in rows if Decimal(r["Amount"]) > 0), Decimal("0"))
    ye_net = ye_debits + ye_credits
    closing = opening_balance + ye_net
    ws3.append(["YE start date", ye_open_date.strftime("%d/%m/%Y")])
    ws3.append(["YE end date", ye_close_date.strftime("%d/%m/%Y")])
    ws3.append(["Opening balance", float(opening_balance)])
    ws3.append(["Total transactions", len(rows)])
    ws3.append(["Total debits (spend)", float(ye_debits)])
    ws3.append(["Total credits (payments+refunds)", float(ye_credits)])
    ws3.append(["Net movement", float(ye_net)])
    ws3.append(["Closing balance (computed)", float(closing)])

    for ws_ in (ws, ws2, ws3):
        for col in ws_.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0 for c in col), default=10)
            ws_.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(path)


# ---------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(description="Convert Halifax 1136 PDFs to Xero CSV")
    p.add_argument("--input-dir", required=True, help="Folder containing the monthly Halifax 1136 PDFs")
    p.add_argument("--output-dir", required=True, help="Where to write the CSV and XLSX")
    p.add_argument("--opening-balance", required=True, type=str, help="YE opening balance (prior YE closing), signed GBP e.g. -1786.33")
    p.add_argument("--ye-year", required=True, type=int, help="YE year, e.g. 2025 for YE 30 Nov 2025")
    p.add_argument("--date-today", required=True, help="YYYY.MM.DD - used in output filenames")
    p.add_argument("--account-code", default="1136", help="Card account code (default 1136)")
    return p.parse_args()


def discover_pdfs(input_dir):
    """Return PDFs in input_dir sorted by name."""
    paths = sorted(glob.glob(os.path.join(input_dir, "*.pdf")))
    if not paths:
        raise FileNotFoundError(f"No .pdf files found in {input_dir}")
    return paths


def main():
    args = parse_args()
    opening_balance = Decimal(args.opening_balance)
    ye_close = date(args.ye_year, 11, 30)
    ye_open = date(args.ye_year - 1, 12, 1)

    pdfs = discover_pdfs(args.input_dir)
    print(f"Found {len(pdfs)} PDF(s) in {args.input_dir}")
    if len(pdfs) != 13:
        print(f"WARNING: expected 13 PDFs (Dec prior-year through Dec YE-year), found {len(pdfs)}")

    stmts = []
    print()
    print(f"{'File':<60}  {'StmtDate':<10}  {'Prev':>10}  {'Pmts':>10}  {'NewChg':>10}  {'NewBal':>10}  {'N':>4}  {'Recon':>10}")
    print("-" * 150)
    problems = []
    for path in pdfs:
        s = parse_statement(path)
        stmts.append(s)
        total = sum((t.amount_gbp for t in s.transactions), Decimal("0"))
        expected = s.payments_received - s.new_charges
        recon = total - expected
        marker = "OK" if recon == 0 else f"DIFF {recon}"
        fname = os.path.basename(path)
        print(f"{fname[:60]:<60}  {s.statement_date}  {s.previous_balance:>10}  {s.payments_received:>10}  {s.new_charges:>10}  {s.new_balance:>10}  {len(s.transactions):>4}  {marker:>10}")
        if recon != 0:
            problems.append((fname, "stmt reconciliation", recon))

    if problems:
        print()
        print("STATEMENT RECONCILIATION FAILURES:")
        for f, kind, v in problems:
            print(f"  {f}: {kind}: {v}")
        sys.exit(1)

    # Build YE rows
    rows = build_rows(stmts, opening_balance, ye_open, ye_close)
    print(f"\nYE {args.ye_year} window: {ye_open} to {ye_close}")
    print(f"Opening balance (input): {opening_balance}")
    print(f"Filtered to {len(rows)} transactions within YE")

    # Verify interior statement boundaries
    print(f"\nStatement-boundary reconciliation:")
    boundary_problems = []
    for r in rows:
        if r["Notes"] == "<- New Balance":
            running = Decimal(r["Balance"])
            stmt_newbal = Decimal(r["_stmt_newbal"])
            # Halifax reports positive owed amount; our running is negative
            if running != -stmt_newbal:
                boundary_problems.append((r["Date"], running, -stmt_newbal))
            print(f"  {r['Date']}  Running={running:>10}  vs Halifax new bal (negated)={-stmt_newbal:>10}  {'OK' if running == -stmt_newbal else 'MISMATCH'}")
    if boundary_problems:
        print("\nBOUNDARY RECONCILIATION FAILURES (running balance vs Halifax New bal):")
        for d, run, exp in boundary_problems:
            print(f"  {d}: running={run}, expected={exp}")
        sys.exit(2)

    closing = opening_balance + sum((Decimal(r["Amount"]) for r in rows), Decimal("0"))
    print(f"\nYE {args.ye_year} closing balance (computed): {closing}")

    # Write outputs
    os.makedirs(args.output_dir, exist_ok=True)
    code = args.account_code
    csv_name = f"{args.date_today} - D.Box_HX-CC-{code}_Xero-import_YE_{args.ye_year}_DRAFT.csv"
    xlsx_name = f"{args.date_today} - D.Box_HX-CC-{code}_Audit-workings_YE_{args.ye_year}_DRAFT.xlsx"
    csv_path = os.path.join(args.output_dir, csv_name)
    xlsx_path = os.path.join(args.output_dir, xlsx_name)
    write_csv(rows, csv_path)
    write_audit_xlsx(rows, stmts, xlsx_path, opening_balance, ye_open, ye_close)
    print(f"\nWrote {csv_path}")
    print(f"Wrote {xlsx_path}")
    print(f"\nDone. Mick to review the XLSX before dropping _DRAFT from the CSV filename.")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
schwab-to-xero skill: Charles Schwab (Corporate ...366) year-end conversion
for Ditty Box Ltd.

Reads the year's Schwab transaction CSV download and produces the figures Mick
keys into Xero as TWO manual cash transactions (Receive Money / Spend Money)
plus the year-end stock revaluation journal. Outputs:

  1. Audit workings XLSX (2 sheets: Transactions, Summary & Schedules)
  2. Posting-schedule Markdown (numbered steps + flags)
  3. Stock-adjustment audit PDF (replica of Mick's handwritten YE working)

This skill does NOT write a Xero CSV. The Schwab account is USD; Xero is
GBP-only, so every USD figure is converted at the single year-end USD/GBP
rate Mick records himself, and the GBP figure is what is keyed into Xero
(the USD is shown in brackets in the line description, matching prior years).

See SKILL.md for the full workflow and field documentation.
"""
import argparse
import csv
import os
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# Parsing helpers
# ============================================================================

def money_to_float(s):
    """Convert '$1,234.56' / '-$0.08' / '' to float (0.0 if blank)."""
    if s is None:
        return 0.0
    s = str(s).replace('$', '').replace(',', '').strip()
    return float(s) if s else 0.0


def parse_schwab_date(s):
    """Schwab dates may be 'MM/DD/YYYY' or 'MM/DD/YYYY as of MM/DD/YYYY'.
    Use the booking (first) date."""
    s = s.split(' as of ')[0].strip()
    return datetime.strptime(s, '%m/%d/%Y')


# ============================================================================
# Classification - Mick's 6-code scheme
# ============================================================================
# Code 1 Div (+)      : Qualified Dividend, Special Qual Div
# Code 2 Sale (+)     : Sell, Cash/Stock Merger (cash leg), Cash In Lieu*
# Code 3 Fees (-)     : ADR Mgmt Fee            -> Xero "SDRT Levy"
# Code 4 Purch (-)    : Buy
# Code 5 WHT (-)      : NRA Tax Adj, Foreign Tax Paid -> Xero "Provision for CT"
# Code 6 Acct Int (+) : Credit Interest, ADR Mgmt Fee Adj (refund)
# Non-cash (no code)  : Stock Merger / Cash/Stock Merger share legs,
#                       Reverse Split, share receipts (Amount blank)
#
# * Cash In Lieu is disposal proceeds of a fractional share. Default treatment
#   is code 2 (Sale). Pass --cil-to-dividends to route it to code 1 instead
#   (matches the one-off YE 2024 treatment where a cash-in-lieu was folded
#   into the dividend line).

CODE_NAMES = {
    1: 'Dividends (+)',
    2: 'Sale of US shares (+)',
    3: 'Fees (-)',
    4: 'Purchase of shares (-)',
    5: 'Withholding Tax (-)',
    6: 'Account Interest (+)',
}


def classify(action, amount, cil_to_dividends=False):
    """Return (code, note). code is None for non-cash rows."""
    a = action.strip()
    if a in ('Qualified Dividend', 'Special Qual Div'):
        return 1, ''
    if a == 'Sell':
        return 2, ''
    if a in ('NRA Tax Adj', 'Foreign Tax Paid'):
        return 5, ''
    if a == 'Buy':
        return 4, ''
    if a in ('Credit Interest', 'ADR Mgmt Fee Adj'):
        return 6, ''
    if a == 'ADR Mgmt Fee':
        return 3, ''
    if a == 'Cash In Lieu':
        return (1, 'cash-in-lieu->Div') if cil_to_dividends else (2, 'cash-in-lieu->Sale')
    if a == 'Cash/Stock Merger':
        if amount != 0:
            return 2, 'merger cash->Sale'
        return None, 'merger share leg (non-cash)'
    if a in ('Stock Merger', 'Reverse Split'):
        return None, 'corporate action (non-cash)'
    # Anything else carrying cash is unexpected -> flag, do not silently drop
    if amount != 0:
        return 'FLAG', 'UNCLASSIFIED cash row - needs manual review'
    return None, 'non-cash'


# ============================================================================
# XLSX writer (with Cowork null-pad workaround)
# ============================================================================

def save_xlsx(wb, final_path):
    """Save to /tmp then cp to final to dodge the Cowork null-pad zip bug."""
    tmp_path = Path(tempfile.gettempdir()) / Path(final_path).name
    wb.save(str(tmp_path))
    shutil.copyfile(str(tmp_path), final_path)


GBPFMT = '#,##0.00;(#,##0.00)'
ARIAL = 'Arial'


def build_workbook(recs, totals, args):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'
    hdr = ['Date', 'Action', 'Symbol', 'Description', 'Code', 'Quantity',
           'Price', 'Fees & Comm', 'Amount (USD)', '1 Div (+$)', '2 Sale (+$)',
           '3 Fees (-$)', '4 Purch (-$)', '5 WHT (-$)', '6 Acct Int (+$)', 'Note']
    HF = Font(name=ARIAL, bold=True, color='FFFFFF')
    HB = PatternFill('solid', start_color='2F5496')
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h)
        cell.font = HF
        cell.fill = HB
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    catcol = {1: 10, 2: 11, 3: 12, 4: 13, 5: 14, 6: 15}
    r0 = 2
    for i, rec in enumerate(recs):
        rr = r0 + i
        ws.cell(rr, 1, rec['dt']).number_format = 'dd/mm/yyyy'
        ws.cell(rr, 2, rec['action'])
        ws.cell(rr, 3, rec['sym'])
        ws.cell(rr, 4, rec['desc'])
        ws.cell(rr, 5, rec['code'] if isinstance(rec['code'], int) else (rec['code'] or ''))
        ws.cell(rr, 6, rec['qty'] if rec['qty'] else None)
        ws.cell(rr, 7, rec['price'] if rec['price'] else None)
        ws.cell(rr, 8, rec['fee'] if rec['fee'] else None)
        ws.cell(rr, 9, rec['amt']).number_format = GBPFMT
        if isinstance(rec['code'], int) and rec['code'] in catcol:
            ws.cell(rr, catcol[rec['code']], rec['amt']).number_format = GBPFMT
        ws.cell(rr, 16, rec['note'])
        for c in range(1, 17):
            ws.cell(rr, c).font = Font(name=ARIAL, size=10)
            ws.cell(rr, c).border = border
    last = r0 + len(recs) - 1
    trow = last + 1
    ws.cell(trow, 4, 'TOTALS (USD)').font = Font(name=ARIAL, bold=True)
    ws.cell(trow, 4).alignment = Alignment(horizontal='right')
    for c in [9, 10, 11, 12, 13, 14, 15]:
        L = get_column_letter(c)
        t = ws.cell(trow, c, f'=SUM({L}{r0}:{L}{last})')
        t.font = Font(name=ARIAL, bold=True)
        t.number_format = GBPFMT
        t.fill = PatternFill('solid', start_color='E8EEF7')
    widths = {1: 11, 2: 18, 3: 9, 4: 42, 5: 6, 6: 9, 7: 9, 8: 11, 9: 13,
              10: 11, 11: 11, 12: 11, 13: 11, 14: 11, 15: 13, 16: 22}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'

    # ---- Summary & Schedules ----
    s = wb.create_sheet('Summary & Schedules')
    B = Font(name=ARIAL, bold=True)
    N = Font(name=ARIAL, size=10)
    TITLE = Font(name=ARIAL, bold=True, size=13, color='2F5496')
    tx = 'Transactions'
    T = str(trow)

    def put(cell, val, font=None, fmt=None, fill=None, align=None):
        c = s[cell]
        c.value = val
        if font:
            c.font = font
        if fmt:
            c.number_format = fmt
        if fill:
            c.fill = PatternFill('solid', start_color=fill)
        if align:
            c.alignment = Alignment(horizontal=align)
        return c

    yelabel = args.ye_date
    put('A1', f'D.Box Schwab USD Account (...366) - YE {yelabel}', TITLE)
    put('A2', 'Year-end USD/GBP rate:', N)
    put('E2', args.usd_rate, Font(name=ARIAL, bold=True, color='0000FF'), '0.000000', fill='FFFF00')

    put('A4', "RECEIVE MONEY  (Xero: Money Rec'd)", B, fill='D9E1F2')
    put('A5', 'Line', B); put('C5', 'USD', B, align='right'); put('D5', 'Rate', B, align='right'); put('E5', 'GBP', B, align='right')
    put('A6', 'Sale of US shares'); put('C6', f"='{tx}'!K{T}", N, GBPFMT); put('D6', '=$E$2'); put('E6', '=C6*D6', N, GBPFMT)
    put('A7', 'Dividends Received'); put('C7', f"='{tx}'!J{T}", N, GBPFMT); put('D7', '=$E$2'); put('E7', '=C7*D7', N, GBPFMT)
    put('A8', "Account Interest Rec'd"); put('C8', f"='{tx}'!O{T}", N, GBPFMT); put('D8', '=$E$2'); put('E8', '=C8*D8', N, GBPFMT)
    put('A9', 'TOTAL RECEIVE', B); put('C9', '=SUM(C6:C8)', B, GBPFMT); put('E9', '=SUM(E6:E8)', B, GBPFMT, fill='E2EFDA')

    put('A11', 'SPEND MONEY  (Xero: SChwab - Costs)', B, fill='D9E1F2')
    put('A12', 'Line', B); put('C12', 'USD', B, align='right'); put('D12', 'Rate', B, align='right'); put('E12', 'GBP', B, align='right')
    put('A13', 'Purchase of Shares (US)'); put('C13', f"=-'{tx}'!M{T}", N, GBPFMT); put('D13', '=$E$2'); put('E13', '=C13*D13', N, GBPFMT)
    put('A14', 'Withholding Tax'); put('C14', f"=-'{tx}'!N{T}", N, GBPFMT); put('D14', '=$E$2'); put('E14', '=C14*D14', N, GBPFMT)
    put('A15', 'Fees / SDRT (ADR mgmt)'); put('C15', f"=-'{tx}'!L{T}", N, GBPFMT); put('D15', '=$E$2'); put('E15', '=C15*D15', N, GBPFMT)
    put('A16', 'TOTAL SPEND', B); put('C16', '=SUM(C13:C15)', B, GBPFMT); put('E16', '=SUM(E13:E15)', B, GBPFMT, fill='FCE4D6')

    put('A18', 'NET CASH MOVEMENT (USD)', B); put('C18', '=C9-C16', B, GBPFMT)
    put('A19', 'Opening cash (prior YE stmt)'); put('C19', args.opening_cash, N, GBPFMT)
    put('A20', 'Closing cash (YE Positions screen)'); put('C20', args.closing_cash, N, GBPFMT)
    put('A21', 'Check: opening + net = closing', B); put('C21', '=C19+C18', B, GBPFMT)
    put('D21', '=IF(ROUND(C21-C20,2)=0,"RECONCILED","CHECK")', B)

    put('A24', 'STOCK REVALUATION JOURNAL  (lower of cost or NRV, whole account)', B, fill='D9E1F2')
    put('A25', 'Cost basis at YE (USD)'); put('C25', args.cost_basis, N, GBPFMT)
    put('A26', 'Market value at YE (USD)'); put('C26', args.market_value, N, GBPFMT)
    put('A27', 'Lower value taken (USD)', B); put('C27', '=MIN(C25:C26)', B, GBPFMT)
    put('A28', 'In GBP at year-end rate', B); put('C28', '=C27*$E$2', B, GBPFMT, fill='E2EFDA')
    put('A29', 'Prior carrying value (GBP)'); put('C29', args.prior_carrying_gbp, N, GBPFMT)
    put('A30', 'Movement in stock value (GBP)', B); put('C30', '=C28-C29', B, GBPFMT)
    put('A32', 'Manual Journal (dated YE):', B); put('C32', 'Dr', B, align='right'); put('E32', 'Cr', B, align='right')
    put('A33', "Dr  637  Val'n of Shares Held"); put('C33', '=IF(C30>0,ABS(C30),0)', N, GBPFMT); put('E33', '=IF(C30<0,ABS(C30),0)', N, GBPFMT)
    put('A34', 'Cr  352  Stock YE'); put('C34', '=IF(C30<0,ABS(C30),0)', N, GBPFMT); put('E34', '=IF(C30>0,ABS(C30),0)', N, GBPFMT)
    put('A36', '(Increase: Dr 637 / Cr 352.  Decrease: Dr 352 / Cr 637.)', Font(name=ARIAL, italic=True, size=9))

    for col, w in {'A': 40, 'B': 6, 'C': 13, 'D': 10, 'E': 13}.items():
        s.column_dimensions[col].width = w
    return wb


# ============================================================================
# Posting note (Markdown)
# ============================================================================

def write_posting_note(path, totals, args, counts, flags, recon_ok):
    r = args.usd_rate
    def g(x):
        return round(x * r, 2)
    lines = []
    lines.append(f"# Schwab USD Account (...366) - YE {args.ye_date} - What to post in Xero")
    lines.append('')
    lines.append(f"Prepared: {args.date_today}. All figures verified in code.")
    lines.append(f"Year-end USD/GBP rate used: {r:.6f}.")
    lines.append('')
    status = 'RECONCILED to the cent' if recon_ok else 'CHECK - DID NOT RECONCILE'
    net = totals['receive'] - totals['spend']
    lines.append(f"Cash reconciliation: opening {args.opening_cash:,.2f} + net {net:,.2f} "
                 f"= {args.opening_cash + net:,.2f} vs closing {args.closing_cash:,.2f}  [{status}]")
    lines.append('')
    lines.append('---')
    lines.append('')
    lines.append('## 1. RECEIVE MONEY (dated YE)')
    lines.append('')
    lines.append(f"1. Sale of US shares                 USD {totals['sale']:>10,.2f}  ->  GBP {g(totals['sale']):>9,.2f}")
    lines.append(f"2. Dividend Income                   USD {totals['div']:>10,.2f}  ->  GBP {g(totals['div']):>9,.2f}")
    lines.append(f"3. Stockbroking Account Interest     USD {totals['acctint']:>10,.2f}  ->  GBP {g(totals['acctint']):>9,.2f}")
    lines.append(f"   TOTAL RECEIVE                     USD {totals['receive']:>10,.2f}  ->  GBP {g(totals['sale'])+g(totals['div'])+g(totals['acctint']):>9,.2f}")
    lines.append('')
    lines.append('## 2. SPEND MONEY (dated YE)')
    lines.append('')
    lines.append(f"1. Purchase of Options/Shares (US)   USD {totals['purch']:>10,.2f}  ->  GBP {g(totals['purch']):>9,.2f}")
    lines.append(f"2. Provision for Corp Tax (withhold) USD {totals['wht']:>10,.2f}  ->  GBP {g(totals['wht']):>9,.2f}")
    lines.append(f"3. SDRT Levy (ADR mgmt fees)         USD {totals['fees']:>10,.2f}  ->  GBP {g(totals['fees']):>9,.2f}")
    lines.append(f"   TOTAL SPEND                       USD {totals['spend']:>10,.2f}  ->  GBP {g(totals['purch'])+g(totals['wht'])+g(totals['fees']):>9,.2f}")
    lines.append('')
    lower = min(args.cost_basis, args.market_value)
    gbp_val = round(lower * r, 2)
    move = round(gbp_val - args.prior_carrying_gbp, 2)
    direction = 'Dr 637 Val\'n of Shares Held / Cr 352 Stock YE' if move > 0 else 'Dr 352 Stock YE / Cr 637 Val\'n of Shares Held'
    lines.append('## 3. STOCK REVALUATION JOURNAL (dated YE)')
    lines.append('')
    lines.append(f"- Cost basis (USD):   {args.cost_basis:,.2f}")
    lines.append(f"- Market value (USD): {args.market_value:,.2f}")
    lines.append(f"- Lower taken (USD):  {lower:,.2f}")
    lines.append(f"- In GBP at {r:.6f}: {gbp_val:,.2f}")
    lines.append(f"- Prior carrying (GBP): {args.prior_carrying_gbp:,.2f}")
    lines.append(f"- Movement (GBP):     {move:,.2f}")
    lines.append(f"- Journal: {direction}  {abs(move):,.2f}")
    lines.append('')
    lines.append('## Classification counts')
    lines.append('')
    for code in sorted(c for c in counts if isinstance(c, int)):
        lines.append(f"- Code {code} {CODE_NAMES[code]}: {counts[code]} rows")
    if 'noncash' in counts:
        lines.append(f"- Non-cash corporate actions (no posting): {counts['noncash']} rows")
    lines.append('')
    if flags:
        lines.append('## FLAGS for Mick')
        lines.append('')
        for f in flags:
            lines.append(f"- {f}")
        lines.append('')
    with open(path, 'w', encoding='ascii', errors='replace') as fh:
        fh.write('\n'.join(lines) + '\n')


# ============================================================================
# Stock-adjustment audit PDF (replica of Mick's handwritten YE working)
# ============================================================================

def prior_ye_label(ye_date):
    """ye_date 'YYYY.MM.DD' -> prior year 'DD.MM.YYYY' and current 'DD.MM.YYYY'."""
    y, m, d = ye_date.split('.')
    cur = f"{d}.{m}.{y}"
    pri = f"{d}.{m}.{int(y) - 1}"
    return pri, cur


def build_stock_pdf(path, args):
    """Typeset audit-trail sheet mirroring Mick's handwritten YE stock
    adjustment working. Skipped with a note if reportlab is unavailable."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError:
        print("NOTE: reportlab not installed - stock-adjustment PDF skipped. "
              "pip install reportlab --break-system-packages", file=sys.stderr)
        return None

    tmp_path = Path(tempfile.gettempdir()) / Path(path).name
    W, H = A4
    c = canvas.Canvas(str(tmp_path), pagesize=A4)

    def text(x, y, s, size=12, font='Helvetica', align='l'):
        c.setFont(font, size)
        if align == 'r':
            c.drawRightString(x, y, s)
        elif align == 'c':
            c.drawCentredString(x, y, s)
        else:
            c.drawString(x, y, s)

    def uline(x1, y, x2, w=0.8):
        c.setLineWidth(w)
        c.line(x1, y, x2, y)

    r = args.usd_rate
    cost, mkt = args.cost_basis, args.market_value
    low = min(cost, mkt)
    g_cur = round(low * r, 2)
    g_pri = args.prior_carrying_gbp
    move = round(g_cur - g_pri, 2)
    word = 'Increased' if move > 0 else 'Decreased'
    pri_lbl, cur_lbl = prior_ye_label(args.ye_date)
    pc, pm, pr = args.prior_cost_basis, args.prior_market_value, args.prior_rate
    have_prior_usd = pc is not None and pm is not None
    p_low = min(pc, pm) if have_prior_usd else None

    def usd(v):
        return f"${v:,.2f}"

    def gbp(v):
        return f"\u00a3{v:,.2f}"  # \u00a3 = pound sign; escaped to keep source ASCII

    left = 28 * mm
    y = H - 28 * mm
    text(W - 28 * mm, y, _pretty_date(args.date_today), size=13, align='r')
    uline(W - 72 * mm, y - 2 * mm, W - 28 * mm)
    y -= 18 * mm

    text(left, y, 'D.Box - Schwab $ Account - Y/E Stock Adjustment', size=15, font='Helvetica-Bold')
    uline(left, y - 2.5 * mm, left + 150 * mm, 1.0)
    uline(left, y - 3.8 * mm, left + 150 * mm, 0.6)
    y -= 20 * mm

    colD, colC, colM, colL = left, left + 45 * mm, left + 85 * mm, left + 125 * mm
    text(colD, y, 'Date', font='Helvetica-Bold'); uline(colD, y - 2 * mm, colD + 18 * mm)
    text(colC, y, 'Cost Basis ($)', font='Helvetica-Bold'); uline(colC, y - 2 * mm, colC + 32 * mm)
    text(colM, y, 'Mkt Value ($)', font='Helvetica-Bold'); uline(colM, y - 2 * mm, colM + 32 * mm)
    text(colL, y, 'Lower Value', font='Helvetica-Bold'); uline(colL, y - 2 * mm, colL + 30 * mm)
    y -= 12 * mm
    text(colD, y, pri_lbl)
    if have_prior_usd:
        text(colC, y, f"{pc:,.2f}"); text(colM, y, f"{pm:,.2f}"); text(colL, y, usd(p_low))
    else:
        text(colL, y, '(see GBP below)')
    y -= 11 * mm
    text(colD, y, cur_lbl); text(colC, y, f"{cost:,.2f}"); text(colM, y, f"{mkt:,.2f}"); text(colL, y, usd(low))
    y -= 18 * mm

    text(left, y, 'In GBP Terms (for Xero)', size=13, font='Helvetica-Bold')
    uline(left, y - 2.5 * mm, left + 58 * mm, 1.0)
    uline(left, y - 3.8 * mm, left + 58 * mm, 0.6)
    y -= 14 * mm
    if have_prior_usd and pr is not None:
        text(left, y, f"{pri_lbl}     {usd(p_low)}  x  {pr:.6f}   =   {gbp(g_pri)}")
    else:
        text(left, y, f"{pri_lbl}     carried at booked value   =   {gbp(g_pri)}")
    y -= 11 * mm
    text(left, y, f"{cur_lbl}     {usd(low)}  x  {r:.6f}   =   {gbp(g_cur)}")
    uline(left + 96 * mm, y - 2.5 * mm, left + 128 * mm, 1.0)
    uline(left + 96 * mm, y - 3.8 * mm, left + 128 * mm, 0.6)
    y -= 18 * mm

    text(left, y, f"Stock Value Y/E {cur_lbl} {word} by:    {gbp(abs(move))}", size=13, font='Helvetica-Bold')
    uline(left + 92 * mm, y - 2.5 * mm, left + 122 * mm, 1.0)
    uline(left + 92 * mm, y - 3.8 * mm, left + 122 * mm, 0.6)
    y -= 20 * mm

    text(left, y, f"Required Manual Journal ({cur_lbl})", size=13, font='Helvetica-Bold')
    y -= 14 * mm
    drx, crx, divx = left + 78 * mm, left + 118 * mm, left + 100 * mm
    text(drx, y, 'Dr', font='Helvetica-Bold', align='r')
    text(crx, y, 'Cr', font='Helvetica-Bold', align='r')
    uline(left + 58 * mm, y - 2 * mm, drx, 0.8)
    uline(divx + 4 * mm, y - 2 * mm, crx, 0.8)
    jtop = y - 3 * mm
    amt = abs(move)
    # Increase: Dr 637 / Cr 352. Decrease: Dr 352 / Cr 637.
    s352_dr, s352_cr = ('-', f"{amt:,.2f}") if move > 0 else (f"{amt:,.2f}", '-')
    s637_dr, s637_cr = (f"{amt:,.2f}", '-') if move > 0 else ('-', f"{amt:,.2f}")
    y -= 12 * mm
    text(left, y, 'Stock Y/E  [352]'); text(drx, y, s352_dr, align='r'); text(crx, y, s352_cr, align='r')
    y -= 13 * mm
    text(left, y, "Val'n of Shares Held  [637]"); text(drx, y, s637_dr, align='r'); text(crx, y, s637_cr, align='r')
    jbot = y - 4 * mm
    c.setLineWidth(0.8)
    c.line(divx, jtop, divx, jbot)
    y -= 22 * mm

    c.setFont('Helvetica-Oblique', 9)
    notes = [
        'Basis: whole-account lower of cost or net realisable value (market), converted to GBP at the year-end USD/GBP rate.',
        f'Prior-year ({pri_lbl}) carried at the booked figure {gbp(g_pri)}. Year-end rates: XE.com mid-market.',
        f'Cost basis and market value at {cur_lbl} from the Schwab Positions screen (cost {cost:,.2f} / market {mkt:,.2f}).',
        f'Prepared by Cedric (Cowork) {args.date_today} as an audit-trail replica of Mick\'s handwritten YE working.',
    ]
    for n in notes:
        c.drawString(left, y, n)
        y -= 4.6 * mm

    c.showPage()
    c.save()
    shutil.copyfile(str(tmp_path), path)
    return path


def _pretty_date(d):
    """'YYYY.MM.DD' -> 'Dth Month YYYY' (e.g. '5th June 2026')."""
    try:
        dt = datetime.strptime(d, '%Y.%m.%d')
    except ValueError:
        return d
    n = dt.day
    suf = 'th' if 11 <= n % 100 <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return f"{n}{suf} {dt.strftime('%B %Y')}"


# ============================================================================
# Main
# ============================================================================

def main():
    ap = argparse.ArgumentParser(description='schwab-to-xero: process Schwab YE transactions')
    ap.add_argument('--csv', required=True, help='Schwab transactions CSV download')
    ap.add_argument('--output-dir', required=True, help='Where to write outputs')
    ap.add_argument('--ye-date', required=True, help='YE date YYYY.MM.DD')
    ap.add_argument('--date-today', required=True, help='Today YYYY.MM.DD (for filenames)')
    ap.add_argument('--usd-rate', required=True, type=float, help='Year-end USD/GBP rate (6dp)')
    ap.add_argument('--cost-basis', required=True, type=float, help='YE cost basis USD (Positions screen)')
    ap.add_argument('--market-value', required=True, type=float, help='YE market value USD (Positions screen)')
    ap.add_argument('--prior-carrying-gbp', required=True, type=float, help='Prior YE stock carrying value GBP (as booked)')
    ap.add_argument('--prior-cost-basis', type=float, default=None, help='Prior YE cost basis USD (for the audit PDF prior row)')
    ap.add_argument('--prior-market-value', type=float, default=None, help='Prior YE market value USD (for the audit PDF prior row)')
    ap.add_argument('--prior-rate', type=float, default=None, help='Prior YE USD/GBP rate (for the audit PDF prior row)')
    ap.add_argument('--opening-cash', required=True, type=float, help='Opening USD cash (prior YE closing)')
    ap.add_argument('--closing-cash', required=True, type=float, help='Closing USD cash (YE Positions screen)')
    ap.add_argument('--cil-to-dividends', action='store_true', help='Route Cash In Lieu to Dividends not Sale')
    args = ap.parse_args()

    with open(args.csv, newline='', encoding='utf-8-sig') as fh:
        rows = list(csv.DictReader(fh))

    recs = []
    cats = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0, 5: 0.0, 6: 0.0}
    counts = {}
    flags = []
    for r in rows:
        amt = money_to_float(r.get('Amount'))
        code, note = classify(r['Action'], amt, args.cil_to_dividends)
        recs.append({
            'dt': parse_schwab_date(r['Date']),
            'action': r['Action'], 'sym': r['Symbol'], 'desc': r['Description'],
            'qty': money_to_float(r.get('Quantity')) or None,
            'price': money_to_float(r.get('Price')) or None,
            'fee': money_to_float(r.get('Fees & Comm')) or None,
            'amt': amt, 'code': code, 'note': note,
        })
        if isinstance(code, int):
            cats[code] += amt
            counts[code] = counts.get(code, 0) + 1
        elif code == 'FLAG':
            flags.append(f"UNCLASSIFIED row {r['Date']} {r['Action']} {r['Symbol']} {amt:,.2f} - needs manual coding")
        else:
            if amt == 0:
                counts['noncash'] = counts.get('noncash', 0) + 1

    recs.sort(key=lambda x: x['dt'])

    totals = {
        'div': cats[1], 'sale': cats[2], 'fees': -cats[3], 'purch': -cats[4],
        'wht': -cats[5], 'acctint': cats[6],
    }
    totals['receive'] = cats[1] + cats[2] + cats[6]
    totals['spend'] = -(cats[3] + cats[4] + cats[5])

    net = round(totals['receive'] - totals['spend'], 2)
    recon_ok = round(args.opening_cash + net - args.closing_cash, 2) == 0

    # Standing flags
    if any(rec['action'] == 'Cash/Stock Merger' and rec['amt'] != 0 for rec in recs):
        flags.append("Merger cash treated as Sale of US shares (disposal proceeds). Confirm.")
    if any(rec['action'] == 'Cash In Lieu' for rec in recs):
        dest = 'Dividends' if args.cil_to_dividends else 'Sale of US shares'
        flags.append(f"Cash In Lieu routed to {dest}. Pass --cil-to-dividends to switch.")
    if min(args.cost_basis, args.market_value) == args.cost_basis:
        flags.append("Lower value is COST this year (market is higher; unrealised gain NOT recognised).")
    else:
        flags.append("Lower value is MARKET this year (market below cost).")

    os.makedirs(args.output_dir, exist_ok=True)
    base = f"{args.date_today} - Schwab Account - YE {args.ye_date}"
    xlsx_path = os.path.join(args.output_dir, base + " Workings and Schedules.xlsx")
    md_path = os.path.join(args.output_dir, base + " Posting Schedule.md")
    pdf_path = os.path.join(args.output_dir, base + " Stock Adjustment (audit trail).pdf")

    wb = build_workbook(recs, totals, args)
    save_xlsx(wb, xlsx_path)
    write_posting_note(md_path, totals, args, counts, flags, recon_ok)
    pdf_written = build_stock_pdf(pdf_path, args)

    # ---- Run summary to stdout ----
    r = args.usd_rate
    print("=== schwab-to-xero run summary ===")
    print(f"Rows: {len(recs)}   Rate: {r:.6f}   YE: {args.ye_date}")
    for code in [1, 2, 3, 4, 5, 6]:
        print(f"  Code {code} {CODE_NAMES[code]:24s} USD {cats[code]:>11.2f}")
    print(f"RECEIVE USD {totals['receive']:.2f}  GBP {totals['receive']*r:.2f}")
    print(f"SPEND   USD {totals['spend']:.2f}  GBP {totals['spend']*r:.2f}")
    print(f"NET     USD {net:.2f}")
    print(f"Cash recon: opening {args.opening_cash:.2f} + net {net:.2f} = "
          f"{args.opening_cash+net:.2f} vs closing {args.closing_cash:.2f}  "
          f"[{'RECONCILED' if recon_ok else 'CHECK'}]")
    lower = min(args.cost_basis, args.market_value)
    move = round(lower*r - args.prior_carrying_gbp, 2)
    print(f"Stock: lower USD {lower:.2f} -> GBP {lower*r:.2f}; prior {args.prior_carrying_gbp:.2f}; movement {move:.2f}")
    if flags:
        print("FLAGS:")
        for f in flags:
            print("  -", f)
    print(f"Outputs:\n  {xlsx_path}\n  {md_path}")
    if pdf_written:
        print(f"  {pdf_path}")

    if not recon_ok:
        print("ERROR: cash did not reconcile - investigate before posting.", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()

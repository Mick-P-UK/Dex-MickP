#!/usr/bin/env python3
"""
PayPal CSV/XLSX to Xero Import Converter
=========================================
Converts a PayPal transaction export (CSV with UTF-8 BOM or XLSX) into:
  1. A 5-column CSV ready for Xero bank statement import
  2. A 20-column audit XLSX with full source detail plus Review
     Annotation column for Mick's manual notes

Usage:
    python3 convert.py \
        --input <paypal-export-path> \
        --output-dir <output-directory> \
        --account-code <xero-code> \
        --account-prefix <filename-prefix> \
        --ye-year <year> \
        --date-today <YYYY.MM.DD>

Example (D.Box PayPal, code 058, YE 2025):
    python3 convert.py \
        --input "/path/to/2026.05.29 - DB-PayPal_01.12.2024-to-30.11.2025_CSV.CSV" \
        --output-dir "C:/Vaults/DB-Accounts-CW/YE_2025.11.30/workings" \
        --account-code "058" \
        --account-prefix "D.Box_PayPal" \
        --ye-year "2025" \
        --date-today "2026.05.29"

Outputs (in --output-dir):
  YYYY.MM.DD - <prefix>-<code>_Xero-import_YE_<year>_DRAFT.csv
  YYYY.MM.DD - <prefix>-<code>_Audit-workings_YE_<year>_DRAFT.xlsx

Exits non-zero if the three-way reconciliation fails.
"""

import argparse
import csv
import os
import shutil
import sys
import tempfile
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------- Known transaction event codes (Ditty Box history) ----------

KNOWN_EVENT_CODES = {
    'T0003',  # Pre-approved Payment Bill User Payment
    'T0007',  # Website Payment
    'T0011',  # Mobile Payment
    'T0200',  # General Currency Conversion
    'T0300',  # Bank deposit to PayPal account
    'T0403',  # User Initiated Withdrawal
    'T0700',  # General Credit Card Deposit
    'T6000',  # Shopping Cart Item (ALWAYS DROPPED)
}


# ---------- Helpers ----------

def parse_decimal(value):
    """Parse a CSV/XLSX value to Decimal, returning Decimal('0') for blank."""
    if value is None or value == '' or (isinstance(value, str) and value.strip() == ''):
        return Decimal('0')
    try:
        return Decimal(str(value).replace(',', '').strip())
    except (InvalidOperation, ValueError):
        return Decimal('0')


def parse_date(value):
    """Parse PayPal date (DD/MM/YYYY) or already-datetime."""
    if isinstance(value, datetime):
        return value.date()
    return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()


# ---------- Source readers ----------

def read_source_csv(path):
    """Read a PayPal CSV export (UTF-8 BOM). Returns list of dicts."""
    with open(path, encoding='utf-8-sig', newline='') as f:
        return list(csv.DictReader(f))


def read_source_xlsx(path):
    """Read a PayPal XLSX export. Returns list of dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(2, ws.max_row + 1):
        d = {}
        for i, h in enumerate(headers):
            v = ws.cell(row=r, column=i + 1).value
            if isinstance(v, datetime):
                d[h] = v.strftime("%d/%m/%Y")
            elif v is None:
                d[h] = ''
            else:
                d[h] = str(v) if not isinstance(v, str) else v
        out.append(d)
    return out


def read_source(path):
    """Auto-detect CSV vs XLSX and read accordingly."""
    if path.lower().endswith('.xlsx'):
        return read_source_xlsx(path)
    return read_source_csv(path)


# ---------- Transformation ----------

def transform(rows):
    """Apply PayPal-to-Xero transformation rules.

    Returns (output_rows, stats) where stats is a dict of counts and
    diagnostics.
    """
    # Stamp source row numbers for traceability
    for i, r in enumerate(rows, start=2):
        r['_source_row'] = i

    # Build TxnID -> Name lookup from parent rows that carry a Name.
    name_lookup = {}
    for r in rows:
        txn_id = (r.get('Transaction ID') or '').strip()
        name = (r.get('Name') or '').strip()
        if txn_id and name and txn_id not in name_lookup:
            name_lookup[txn_id] = name

    # For each GBP General Currency Conversion (T0200), find the matching
    # USD T0200 row sharing the same Reference Txn ID.
    usd_conversion_by_ref = {}
    for r in rows:
        if (r.get('Transaction Event Code') == 'T0200'
                and r.get('Currency') == 'USD'):
            ref = (r.get('Reference Txn ID') or '').strip()
            if ref:
                usd_conversion_by_ref[ref] = parse_decimal(r.get('Gross'))

    # Counters and diagnostics
    stats = {
        'source_rows': len(rows),
        'dropped_t6000': 0,
        'dropped_usd': 0,
        'fee_splits': 0,
        'blank_payee_rows': [],
        'unseen_event_codes': set(),
    }

    output = []
    for r in rows:
        evt = (r.get('Transaction Event Code') or '').strip()
        cur = (r.get('Currency') or '').strip()

        if evt and evt not in KNOWN_EVENT_CODES:
            stats['unseen_event_codes'].add(evt)

        if evt == 'T6000':
            stats['dropped_t6000'] += 1
            continue
        if cur != 'GBP':
            stats['dropped_usd'] += 1
            continue

        # Payee Name (carry forward from parent if blank).
        name = (r.get('Name') or '').strip()
        if not name:
            ref = (r.get('Reference Txn ID') or '').strip()
            if ref and ref in name_lookup:
                name = name_lookup[ref]

        # Description (enrich for currency conversion).
        desc = (r.get('Type') or '').strip()
        if evt == 'T0200':
            ref = (r.get('Reference Txn ID') or '').strip()
            usd_amt = usd_conversion_by_ref.get(ref)
            if usd_amt is not None and usd_amt != 0:
                desc = f"General Currency Conversion (${abs(usd_amt):.2f})"

        gross = parse_decimal(r.get('Gross'))
        fee = parse_decimal(r.get('Fee'))
        net = parse_decimal(r.get('Net'))

        main_row = {
            'Date': parse_date(r.get('Date')),
            'Amount': gross,
            'Payee': name,
            'Description': desc,
            'Reference': (r.get('Transaction ID') or '').strip(),
            'Status': r.get('Status', ''),
            'Currency': cur,
            'Gross': gross,
            'Fee': fee,
            'Net': net,
            'Balance': parse_decimal(r.get('Balance')),
            'Transaction ID': r.get('Transaction ID', ''),
            'Reference Txn ID': r.get('Reference Txn ID', ''),
            'Event Code': evt,
            'Balance Impact': r.get('Balance Impact', ''),
            'From Email': r.get('From Email Address', ''),
            'To Email': r.get('To Email Address', ''),
            'Note': r.get('Note', ''),
            'Source Row': r.get('_source_row', ''),
            'Review Annotation': '',
        }
        output.append(main_row)

        if not name:
            stats['blank_payee_rows'].append((main_row['Date'], gross, desc))

        if fee != 0:
            fee_row = dict(main_row)
            fee_row['Amount'] = fee
            fee_row['Description'] = 'PayPal Fee'
            fee_row['Gross'] = fee
            fee_row['Net'] = Decimal('0')
            fee_row['Balance'] = Decimal('0')
            fee_row['Event Code'] = 'FEE-SPLIT'
            fee_row['Balance Impact'] = 'Debit'
            fee_row['Note'] = ''
            output.append(fee_row)
            stats['fee_splits'] += 1

    # Stable sort by date only - preserves PayPal source order within date.
    output.sort(key=lambda x: x['Date'])
    stats['output_rows'] = len(output)
    return output, stats


# ---------- Reconciliation ----------

def reconcile(source_rows, output_rows):
    """Compute the three-way reconciliation."""
    src_total = Decimal('0')
    for r in source_rows:
        if (r.get('Currency') == 'GBP'
                and r.get('Transaction Event Code') != 'T6000'):
            g = parse_decimal(r.get('Gross'))
            f = parse_decimal(r.get('Fee'))
            src_total += g + f

    last_gbp_bal = Decimal('0')
    for r in source_rows:
        if (r.get('Currency') == 'GBP'
                and r.get('Balance') and str(r.get('Balance')).strip()):
            last_gbp_bal = parse_decimal(r.get('Balance'))

    out_total = sum((r['Amount'] for r in output_rows), Decimal('0'))

    return {
        'source_gbp_movement': src_total,
        'paypal_final_balance': last_gbp_bal,
        'output_amount_total': out_total,
        'ties': src_total == last_gbp_bal == out_total,
    }


# ---------- Writers ----------

def write_xero_csv(output_rows, out_path):
    """Write the Xero bank statement import CSV (5 columns)."""
    # Write to a temp file first to avoid Cowork null-padding on Windows mounts.
    tmp = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv',
                                       encoding='utf-8', newline='')
    try:
        w = csv.writer(tmp)
        w.writerow(['*Date', '*Amount', 'Payee', 'Description', 'Reference'])
        for r in output_rows:
            w.writerow([
                r['Date'].strftime('%d/%m/%Y'),
                f"{r['Amount']:.2f}",
                r['Payee'],
                r['Description'],
                r['Reference'],
            ])
        tmp.close()
        shutil.copy(tmp.name, out_path)
    finally:
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)


def write_audit_xlsx(output_rows, out_path, source_path, account_code, recon):
    """Write the audit XLSX (20 columns plus header block + total row)."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PayPal Audit'

    title_font = Font(bold=True, size=12, color='FFFFFF')
    title_fill = PatternFill('solid', fgColor='305496')
    header_font = Font(bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    review_fill = PatternFill('solid', fgColor='FFF2CC')
    border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )

    ws['A1'] = (f"Ditty Box Ltd - PayPal (Xero {account_code}) - "
                f"Audit Working")
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws.merge_cells('A1:T1')
    ws['A2'] = f"Source: {Path(source_path).name}"
    ws['A3'] = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} "
                f"by paypal-to-xero skill")
    ws['A4'] = ("Transformation: USD rows and Shopping Cart Items (T6000) "
                "dropped. Gross/Fee splits expanded into two rows. Payee "
                "carried forward from parent via Reference Txn ID. General "
                "Currency Conversion descriptions enriched with USD amount. "
                "Review Annotation column (yellow) is for Mick's manual "
                "notes before final Xero import (e.g. payee for User "
                "Initiated Withdrawals, contextual notes).")
    ws.row_dimensions[4].height = 45
    ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A4:T4')

    ws['A5'] = (f"Reconciliation: Source GBP movement = {recon['source_gbp_movement']:.2f}, "
                f"PayPal final balance = {recon['paypal_final_balance']:.2f}, "
                f"Output total = {recon['output_amount_total']:.2f}, "
                f"Ties: {'YES' if recon['ties'] else 'NO - INVESTIGATE'}")
    ws['A5'].font = Font(italic=True)
    ws.merge_cells('A5:T5')

    columns = [
        'Date', 'Amount', 'Payee', 'Description', 'Review Annotation',
        'Reference', 'Status', 'Currency', 'Gross', 'Fee', 'Net', 'Balance',
        'Transaction ID', 'Reference Txn ID', 'Event Code', 'Balance Impact',
        'From Email', 'To Email', 'Note', 'Source Row',
    ]
    header_row = 7
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=i, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    for ri, r in enumerate(output_rows, start=header_row + 1):
        for ci, col in enumerate(columns, start=1):
            v = r.get(col, '')
            if col == 'Date':
                cell = ws.cell(row=ri, column=ci, value=r['Date'])
                cell.number_format = 'dd/mm/yyyy'
            elif col in ('Amount', 'Gross', 'Fee', 'Net', 'Balance'):
                vv = r.get(col, Decimal('0'))
                val = float(vv) if vv != '' else None
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
            else:
                cell = ws.cell(row=ri, column=ci, value=v if v != '' else None)
            cell.border = border
            if col == 'Review Annotation':
                cell.fill = review_fill

    total_row = header_row + 1 + len(output_rows)
    label = ws.cell(row=total_row, column=1, value='TOTAL')
    label.font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=2,
                         value=f"=SUM(B{header_row + 1}:B{total_row - 1})")
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00;[Red]-#,##0.00'
    total_cell.border = Border(top=Side(style='double'))

    widths = {
        'A': 11, 'B': 11, 'C': 28, 'D': 38, 'E': 26,
        'F': 22, 'G': 11, 'H': 9, 'I': 10, 'J': 9, 'K': 10, 'L': 10,
        'M': 22, 'N': 22, 'O': 11, 'P': 13,
        'Q': 28, 'R': 28, 'S': 18, 'T': 11,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f'A{header_row + 1}'

    wb.save(tmp.name)
    shutil.copy(tmp.name, out_path)
    os.unlink(tmp.name)


# ---------- CLI ----------

def parse_args():
    p = argparse.ArgumentParser(
        description="Convert PayPal CSV/XLSX export to Xero import format")
    p.add_argument('--input', required=True,
                   help='Path to PayPal export (CSV or XLSX)')
    p.add_argument('--output-dir', required=True,
                   help='Directory for output files')
    p.add_argument('--account-code', required=True,
                   help='Xero bank code (e.g. 058 for D.Box PayPal)')
    p.add_argument('--account-prefix', required=True,
                   help='Filename prefix (e.g. D.Box_PayPal or M_PayPal)')
    p.add_argument('--ye-year', required=True,
                   help='YE year (e.g. 2025 for YE 30 Nov 2025)')
    p.add_argument('--date-today', required=True,
                   help='Date today in YYYY.MM.DD form for filenames')
    return p.parse_args()


def main():
    args = parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: input file not found: {args.input}")
        sys.exit(1)

    os.makedirs(args.output_dir, exist_ok=True)

    source_rows = read_source(args.input)
    output_rows, stats = transform(source_rows)
    recon = reconcile(source_rows, output_rows)

    base = (f"{args.date_today} - {args.account_prefix}-{args.account_code}"
            f"_{{kind}}_YE_{args.ye_year}_DRAFT.{{ext}}")
    csv_path = os.path.join(args.output_dir,
                             base.format(kind='Xero-import', ext='csv'))
    xlsx_path = os.path.join(args.output_dir,
                             base.format(kind='Audit-workings', ext='xlsx'))

    write_xero_csv(output_rows, csv_path)
    write_audit_xlsx(output_rows, xlsx_path, args.input,
                     args.account_code, recon)

    # Report
    print("=" * 72)
    print(f"PayPal-to-Xero conversion complete")
    print("=" * 72)
    print(f"Source: {args.input}")
    print(f"Account: {args.account_prefix} (Xero {args.account_code})")
    print(f"YE Year: {args.ye_year}")
    print()
    print(f"Source rows: {stats['source_rows']}")
    print(f"  Dropped (T6000 Shopping Cart Items): {stats['dropped_t6000']}")
    print(f"  Dropped (USD currency rows): {stats['dropped_usd']}")
    print(f"Output rows: {stats['output_rows']}")
    print(f"  Of which from fee splits: {stats['fee_splits']}")
    print()
    print("Three-way reconciliation:")
    print(f"  Source GBP non-T6000 net movement: {recon['source_gbp_movement']:>12.2f}")
    print(f"  Source PayPal final balance:       {recon['paypal_final_balance']:>12.2f}")
    print(f"  Output Amount column total:        {recon['output_amount_total']:>12.2f}")
    if recon['ties']:
        print(f"  RESULT: TIES - all three figures agree to the penny")
    else:
        print(f"  RESULT: DOES NOT TIE - investigate before importing")
    print()
    if stats['blank_payee_rows']:
        print(f"Blank Payee rows needing Mick's manual fill ({len(stats['blank_payee_rows'])}):")
        for d, amt, desc in stats['blank_payee_rows']:
            print(f"  {d.strftime('%d/%m/%Y')}  {amt:>10.2f}  {desc}")
        print()
    if stats['unseen_event_codes']:
        print(f"WARNING - unseen Transaction Event Codes: {sorted(stats['unseen_event_codes'])}")
        print("  These were treated as normal GBP rows but may need investigation.")
        print()
    print(f"Wrote: {csv_path}")
    print(f"Wrote: {xlsx_path}")

    sys.exit(0 if recon['ties'] else 2)


if __name__ == '__main__':
    main()

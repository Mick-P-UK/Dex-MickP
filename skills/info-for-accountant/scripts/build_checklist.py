#!/usr/bin/env python3
"""
Build the Ditty Box Ltd accountant-pack checklist spreadsheet.

Usage:
    python build_checklist.py items.json output.xlsx "YE 30.11.2025" "2026.07.26"

items.json is a list of objects:
    [
      {"ref": "J01",
       "item": "NatWest Current Acc - statement showing opening balance ...",
       "status": "Copied",            # "Copied" or "OUTSTANDING"
       "file": "2026.05.27 - DBox_NW-Curr-Acc_Statement-0708_...pdf",
       "folder": "Nat West - Curr Acc (GBP)",
       "note": ""},
      ...
    ]

Writes an Arial-formatted, print-ready sheet with Mick's standard footer.
ASCII only. Run the xlsx skill's recalc.py afterwards and confirm zero errors.
"""

import json
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Mick's house standard - see CLAUDE.md "Spreadsheet Print Footer"
FOOTER = "(&Z&F - Printed: &D at &T)"

ARIAL = "Arial"
HEADERS = [
    "Ref",
    "Item",
    "Status",
    "File name in 'For sending 2 Jade'",
    "Source sub-folder",
    "Notes / action needed",
]
WIDTHS = [8, 52, 14, 62, 34, 58]
HEADER_ROW = 4


def apply_house_footer(workbook):
    """Standard footer plus print setup on every worksheet."""
    for ws in workbook.worksheets:
        ws.oddFooter.left.text = FOOTER
        ws.evenFooter.left.text = FOOTER
        ws.firstFooter.left.text = FOOTER
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "{0}:{0}".format(HEADER_ROW)


def build(items, out_path, ye_label, date_label):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Info for Accountant"

    ws["A1"] = "D.Box Accounts - Info for the accountant (checklist)"
    ws["A1"].font = Font(name=ARIAL, size=14, bold=True)
    ws["E1"] = ye_label
    ws["E1"].font = Font(name=ARIAL, size=14, bold=True)
    ws["A2"] = (
        "Prepared {0}. Copies taken from the year folder - "
        "all originals left in place.".format(date_label)
    )
    ws["A2"].font = Font(name=ARIAL, size=10, italic=True)

    for i, head in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=i, value=head)
        cell.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = PatternFill("solid", fgColor="E2EFDA")
    amber = PatternFill("solid", fgColor="FFF2CC")

    row = HEADER_ROW + 1
    for entry in items:
        status = entry.get("status", "OUTSTANDING")
        values = [
            entry.get("ref", ""),
            entry.get("item", ""),
            status,
            entry.get("file", ""),
            entry.get("folder", ""),
            entry.get("note", ""),
        ]
        for i, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=i, value=value)
            cell.font = Font(name=ARIAL, size=10, bold=(i == 1))
            cell.alignment = Alignment(
                vertical="top", wrap_text=(i in (2, 4, 5, 6))
            )
            cell.border = border
            cell.fill = green if status == "Copied" else amber
        row += 1

    first, last = HEADER_ROW + 1, row - 1
    totals = [
        ("Items copied", '=COUNTIF(C{0}:C{1},"Copied")'.format(first, last)),
        ("Items outstanding", '=COUNTIF(C{0}:C{1},"OUTSTANDING")'.format(first, last)),
        ("Total items", "=COUNTA(A{0}:A{1})".format(first, last)),
    ]
    for offset, (label, formula) in enumerate(totals):
        target = row + 1 + offset
        ws.cell(row=target, column=2, value=label).font = Font(
            name=ARIAL, size=10, bold=True
        )
        ws.cell(row=target, column=3, value=formula).font = Font(
            name=ARIAL, size=10, bold=True
        )

    for col, width in zip("ABCDEF", WIDTHS):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A{0}".format(HEADER_ROW + 1)
    ws.row_dimensions[HEADER_ROW].height = 30

    apply_house_footer(wb)
    wb.save(out_path)
    return out_path


def main():
    if len(sys.argv) != 5:
        print(__doc__)
        return 1
    items_path, out_path, ye_label, date_label = sys.argv[1:5]
    with open(items_path, "r", encoding="utf-8") as handle:
        items = json.load(handle)
    build(items, out_path, ye_label, date_label)
    copied = sum(1 for i in items if i.get("status") == "Copied")
    print(
        "Written {0} - {1} items, {2} copied, {3} outstanding".format(
            out_path, len(items), copied, len(items) - copied
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())

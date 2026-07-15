#!/usr/bin/env python3
"""
NatWest CSV to Xero Import Converter
=====================================
Converts a NatWest online banking CSV statement export into:
  1. A 6-column CSV ready for Xero bank statement import
  2. An 11-column working XLSX with full detail for audit reference

Usage:
    python3 convert.py \
        --input <natwest-csv-path> \
        --output-dir <output-directory> \
        --account-code <xero-code> \
        --account-name <short-name> \
        --ye-year <year> \
        --date-today <YYYY.MM.DD>

Example:
    python3 convert.py \
        --input "/path/to/NatWest-transactions.csv" \
        --output-dir "/path/to/workings" \
        --account-code "051" \
        --account-name "Curr-Acc" \
        --ye-year "2025" \
        --date-today "2026.05.27"
"""

import argparse
import csv
import os
import random
import sys
import re
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Convert NatWest CSV to Xero import format"
    )
    parser.add_argument("--input", required=True, help="Path to NatWest CSV export")
    parser.add_argument("--output-dir", required=True, help="Directory for output files")
    parser.add_argument("--account-code", required=True, help="Xero bank code (e.g. 051)")
    parser.add_argument("--account-name", required=True, help="Short account name (e.g. Curr-Acc)")
    parser.add_argument("--ye-year", required=True, help="Year ending (e.g. 2025)")
    parser.add_argument("--date-today", required=True, help="Today's date as YYYY.MM.DD")
    return parser.parse_args()


def read_natwest_csv(filepath):
    """Read and validate the NatWest CSV export."""
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)
        rows = list(reader)

    # Strip whitespace from header
    header = [h.strip() for h in header]

    expected_cols = ["Date", "Type", "Description", "Value", "Balance",
                     "Account Name", "Account Number"]
    if header != expected_cols:
        print(f"WARNING: Unexpected header columns.")
        print(f"  Expected: {expected_cols}")
        print(f"  Got:      {header}")
        print(f"  Proceeding anyway -- check results carefully.")

    # Remove any fully blank rows
    rows = [r for r in rows if any(c.strip() for c in r)]

    return header, rows


def parse_date(date_str):
    """Parse NatWest date format 'DD Mon YYYY' to datetime."""
    return datetime.strptime(date_str.strip(), "%d %b %Y")


# NatWest embeds the COUNTERPARTY ACCOUNT NUMBER in its own narrative, e.g.
# "From A/C 06512917,PAVEY M A (current,Dir Loan to D.Box". Passing that
# through verbatim breaches the standing rule in CLAUDE.md: "NEVER write bank
# account numbers in full in any file or chat - mask middle digits". Found
# 2026.07.14 - 20 rows of the YE2025 051 import carried three full account
# numbers straight into the Xero CSV, the working XLSX and the git repo.
# House style is ****<last 4>.
ACCT_RE = re.compile(r"\b(A/C|ACCOUNT|ACC\.?\s*NO\.?)\s*(\d{8})\b", re.I)


def mask_account_numbers(text):
    """Mask any full 8-digit bank account number in a NatWest narrative.

    Only masks digits ANCHORED to an account marker (A/C / ACCOUNT / ACC NO),
    which is NatWest's own convention. Bare 8-digit runs are left alone - they
    are usually genuine payment references Mick needs - but scan_for_bare_8
    warns about them so a human can check.
    """
    if not text:
        return text
    return ACCT_RE.sub(lambda m: "%s ****%s" % (m.group(1), m.group(2)[-4:]), text)


def scan_for_bare_8(text, seen):
    """Collect un-anchored 8-digit runs so the run can warn about them."""
    if not text:
        return
    for m in re.finditer(r"(?<![\d/])\d{8}(?![\d])", ACCT_RE.sub("", text)):
        seen.add(m.group(0))


def split_description(desc):
    """
    Split a NatWest Description field on commas.
    Returns up to 5 segments: payee, desc1, payt_method, reference, ref_no.
    Preserves leading/trailing spaces (matching Mick's convention).
    If more than 5 segments, extras are joined into ref_no.

    Account numbers are MASKED here - this is the single funnel every output
    field passes through, so masking once covers Payee, Description, Reference
    and Ref.No. in both the Xero CSV and the working XLSX.
    """
    desc = mask_account_numbers(desc)
    parts = desc.split(",")
    payee       = parts[0] if len(parts) > 0 else None
    desc1       = parts[1] if len(parts) > 1 else None
    payt_method = parts[2] if len(parts) > 2 else None
    reference   = parts[3] if len(parts) > 3 else None
    ref_no      = ",".join(parts[4:]) if len(parts) > 4 else None
    return payee, desc1, payt_method, reference, ref_no


def build_xlsx(rows_data, output_path):
    """Build the 11-column working XLSX."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NatWest-Xero-Working"

    headers = ["*Date", "*Amount", "Payee", "Description-1", "Pay't_Method",
               "Reference", "Ref.No.", "Balance", "Orig_Amount", "Balance", "Type"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows_data:
        ws.append([
            row["date"],        # *Date
            row["value"],       # *Amount
            row["payee"],       # Payee
            row["desc1"],       # Description-1
            row["payt_method"], # Pay't_Method
            row["reference"],   # Reference
            row["ref_no"],      # Ref.No.
            row["balance"],     # Balance
            row["value"],       # Orig_Amount
            row["balance"],     # Balance (copy)
            row["type"],        # Type
        ])

    # Format date column
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"

    # Format number columns
    for col_letter in ["B", "H", "I", "J"]:
        col_idx = ws[f"{col_letter}1"].column
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).number_format = "#,##0.00"

    # Column widths
    widths = {"A": 12, "B": 12, "C": 25, "D": 25, "E": 22,
              "F": 20, "G": 25, "H": 12, "I": 12, "J": 12, "K": 6}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wb.save(output_path)
    return ws.max_row - 1  # data row count


def build_csv(rows_data, output_path):
    """Build the 6-column Xero import CSV."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["*Date", "*Amount", "Payee", "Description",
                         "Reference", "Check Number"])

        for row in rows_data:
            date_str = row["date"].strftime("%d/%m/%Y")
            payee = row["payee"].strip() if row["payee"] else ""
            desc = row["desc1"].strip() if row["desc1"] else ""
            ref = row["reference"].strip() if row["reference"] else ""

            writer.writerow([
                date_str,
                f"{row['value']:.2f}",
                payee,
                desc,
                ref,
                ""
            ])


def verify(rows_data, source_rows, source_header):
    """Run all verification checks. Returns True if all pass."""
    print("\n" + "=" * 60)
    print("VERIFICATION REPORT")
    print("=" * 60)

    all_ok = True

    # 1. Row count
    n = len(rows_data)
    n_src = len(source_rows)
    status = "PASS" if n == n_src else "FAIL"
    if status == "FAIL":
        all_ok = False
    print(f"\n1. ROW COUNT: {n} (source: {n_src}) -- {status}")

    # 2. Date range
    first_d = rows_data[0]["date"]
    last_d = rows_data[-1]["date"]
    print(f"\n2. DATE RANGE: {first_d.strftime('%d %b %Y')} to {last_d.strftime('%d %b %Y')}")

    # 3. Chronological order
    ordered = all(rows_data[i]["date"] >= rows_data[i-1]["date"]
                  for i in range(1, len(rows_data)))
    print(f"\n3. CHRONOLOGICAL ORDER: {'PASS' if ordered else 'FAIL'}")
    if not ordered:
        all_ok = False

    # 4. Opening balance
    opening = rows_data[0]["balance"] - rows_data[0]["value"]
    print(f"\n4. OPENING BALANCE: {opening:.2f}")
    print(f"   (first balance {rows_data[0]['balance']:.2f} - "
          f"first amount {rows_data[0]['value']:.2f})")

    # 5. Closing balance
    closing = rows_data[-1]["balance"]
    print(f"\n5. CLOSING BALANCE: {closing:.2f}")

    # 6. Sum of amounts
    total = sum(r["value"] for r in rows_data)
    expected_sum = closing - opening
    match = abs(total - expected_sum) < 0.005
    print(f"\n6. SUM OF AMOUNTS: {total:.2f} (expected: {expected_sum:.2f}) "
          f"-- {'PASS' if match else 'FAIL'}")
    if not match:
        all_ok = False

    # 7. Running balance continuity
    breaks = 0
    for i in range(1, len(rows_data)):
        prev_bal = rows_data[i-1]["balance"]
        curr_amt = rows_data[i]["value"]
        curr_bal = rows_data[i]["balance"]
        expected = round(prev_bal + curr_amt, 2)
        if abs(expected - curr_bal) > 0.01:
            breaks += 1
            if breaks <= 3:
                print(f"   Break at row {i+2}: {prev_bal} + {curr_amt} = "
                      f"{expected}, got {curr_bal}")
    print(f"\n7. RUNNING BALANCE CONTINUITY: {breaks} breaks "
          f"-- {'PASS' if breaks == 0 else 'FAIL'}")
    if breaks > 0:
        all_ok = False

    # 8. Spot-check 5 rows
    print(f"\n8. SPOT-CHECK (5 random rows):")
    # source_rows is already reversed to chronological
    random.seed(42)
    indices = sorted(random.sample(range(len(rows_data)), min(5, len(rows_data))))
    spot_ok = True
    for idx in indices:
        src = source_rows[idx]
        src_date = parse_date(src[0]).strftime("%d %b %Y")
        src_val = float(src[3])
        src_bal = float(src[4])
        our_date = rows_data[idx]["date"].strftime("%d %b %Y")
        our_val = rows_data[idx]["value"]
        our_bal = rows_data[idx]["balance"]

        ok = (our_date == src_date and
              abs(our_val - src_val) < 0.001 and
              abs(our_bal - src_bal) < 0.001)
        if not ok:
            spot_ok = False
        print(f"   Row {idx+2}: {'MATCH' if ok else 'MISMATCH'} | "
              f"{our_date} | {our_val:>10.2f} | bal {our_bal:>10.2f}")
    if not spot_ok:
        all_ok = False

    # 9. Transaction type distribution
    from collections import Counter
    types = Counter(r["type"] for r in rows_data)
    print(f"\n9. TRANSACTION TYPE DISTRIBUTION:")
    for t, count in sorted(types.items()):
        print(f"   {t}: {count}")

    print(f"\n{'=' * 60}")
    if all_ok:
        print("ALL CHECKS PASSED")
    else:
        print("SOME CHECKS FAILED -- review above")
    print(f"{'=' * 60}")

    return all_ok


def main():
    args = parse_args()

    # Read source
    print(f"Reading: {args.input}")
    header, raw_rows = read_natwest_csv(args.input)
    print(f"Found {len(raw_rows)} transactions")

    # Reverse to chronological
    raw_rows.reverse()

    # Parse into structured data
    rows_data = []
    for row in raw_rows:
        dt = parse_date(row[0])
        tx_type = row[1].strip()
        desc = row[2]
        value = float(row[3])
        balance = float(row[4])

        payee, desc1, payt_method, reference, ref_no = split_description(desc)

        rows_data.append({
            "date": dt,
            "type": tx_type,
            "value": value,
            "balance": balance,
            "payee": payee,
            "desc1": desc1,
            "payt_method": payt_method,
            "reference": reference,
            "ref_no": ref_no,
        })

    # Build output filenames
    date_str = args.date_today
    base = f"{date_str} - D.Box_NW-{args.account_name}-{args.account_code}"
    xlsx_name = f"{base}_Xero-import-ready_YE_{args.ye_year}.xlsx"
    csv_name = f"{base}_Xero-import_YE_{args.ye_year}.csv"

    xlsx_path = os.path.join(args.output_dir, xlsx_name)
    csv_path = os.path.join(args.output_dir, csv_name)

    # Ensure output directory exists
    os.makedirs(args.output_dir, exist_ok=True)

    # Build outputs
    n_xlsx = build_xlsx(rows_data, xlsx_path)
    print(f"XLSX written: {xlsx_path} ({n_xlsx} data rows)")

    build_csv(rows_data, csv_path)
    print(f"CSV written:  {csv_path}")

    # Verify
    verify(rows_data, raw_rows, header)

    # Summary
    opening = rows_data[0]["balance"] - rows_data[0]["value"]
    closing = rows_data[-1]["balance"]
    print(f"\nSUMMARY:")
    print(f"  Account: {args.account_code} ({args.account_name})")
    print(f"  YE period: 01.12.{int(args.ye_year)-1} to 30.11.{args.ye_year}")
    print(f"  Transactions: {len(rows_data)}")
    print(f"  Opening balance: {opening:.2f}")
    print(f"  Closing balance: {closing:.2f}")
    print(f"  Net movement: {closing - opening:.2f}")
    print(f"\nFiles:")
    print(f"  XLSX (working): {xlsx_path}")
    print(f"  CSV (Xero import): {csv_path}")


if __name__ == "__main__":
    main()

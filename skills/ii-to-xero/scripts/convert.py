#!/usr/bin/env python3
"""
ii-to-xero skill: Interactive Investor year-end conversion.

Consolidates the parser, XLSX extender, and audit builder into a single
CLI. Reads the year's ii transaction CSV, extends the master tracking
spreadsheet with a new YE block, and produces three outputs:
  1. Extended tracking spreadsheet XLSX
  2. Audit workings XLSX (3 sheets)
  3. Workings note Markdown

See SKILL.md for the workflow and field-level documentation.
"""
import argparse
import csv
import json
import os
import re
import shutil
import sys
import tempfile
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================================
# Parsing helpers
# ============================================================================

def money_to_float(s):
    """Convert 'GBP1,234.56' or '0.16' or 'n/a' or '' to float or None."""
    if s is None:
        return None
    s = str(s).strip()
    if not s or s == 'n/a':
        return None
    # Strip currency symbol and thousands commas
    s = s.replace('£', '').replace(',', '').strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_ddmmyyyy(s):
    return datetime.strptime(s.strip(), '%d/%m/%Y').date()


def fmt_yyyymmdd(d):
    return d.strftime('%Y.%m.%d')


def classify_row(row):
    desc = row.get('Description', '')
    sym = row.get('Symbol', '')
    debit = row.get('Debit_f')
    credit = row.get('Credit_f')
    if 'GROSS INTEREST' in desc:
        return 'INTEREST'
    if desc.startswith('Div '):
        return 'DIVIDEND'
    if desc.startswith('PAYMENT'):
        return 'WITHDRAWAL' if debit else 'DEPOSIT'
    if sym and sym != 'n/a':
        return 'BUY' if debit else 'SELL'
    if re.match(r'^(\d+)\s', desc):
        return 'BUY' if debit else 'SELL'
    return 'OTHER'


def parse_csv(csv_path):
    """Parse the ii CSV; returns list of dicts with parsed fields."""
    rows = []
    with open(csv_path, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        headers = next(reader)
        # Strip stray BOM characters that sometimes stack
        headers = [h.lstrip('﻿') for h in headers]
        expected_headers = ['Date', 'Settlement Date', 'Symbol', 'Sedol',
                            'Quantity', 'Price', 'Description', 'Reference',
                            'Debit', 'Credit', 'Running Balance']
        if headers[:11] != expected_headers:
            print(f"WARN: header mismatch.\n  Got:      {headers[:11]}\n  Expected: {expected_headers}",
                  file=sys.stderr)
        for csv_idx, row in enumerate(reader):
            if not row or all(not c.strip() for c in row):
                continue
            d = dict(zip(headers, row))
            d['csv_idx'] = csv_idx
            d['Date_parsed'] = parse_ddmmyyyy(d['Date'])
            d['Debit_f'] = money_to_float(d.get('Debit'))
            d['Credit_f'] = money_to_float(d.get('Credit'))
            d['Running Balance_f'] = money_to_float(d.get('Running Balance'))
            d['Quantity_f'] = money_to_float(d.get('Quantity'))
            d['Price_f'] = money_to_float(d.get('Price'))
            d['Category'] = classify_row(d)
            rows.append(d)
    # Sort chronological ascending; same date -> reverse CSV order to recover
    # intraday sequence (ii lists reverse-chron within a day)
    rows.sort(key=lambda r: (r['Date_parsed'], -r['csv_idx']))
    return rows


# ============================================================================
# Ticker -> friendly name lookup
# ============================================================================

TICKER_PATTERN = re.compile(r'\(([A-Z0-9]+)\)')
DOLLAR_TAIL_PATTERN = re.compile(r'\s*\$[\d.]+\s*$')

def build_ticker_lookup(existing_xlsx):
    """Scan the existing tracking spreadsheet's Detail column for friendly names."""
    wb = openpyxl.load_workbook(existing_xlsx, data_only=True)
    ws = wb["Sheet1"]
    lookup = {}
    for r in range(1, ws.max_row + 1):
        detail = ws.cell(row=r, column=2).value
        if not detail or not isinstance(detail, str):
            continue
        m = TICKER_PATTERN.search(detail)
        if not m:
            continue
        ticker = m.group(1)
        if ticker.lower() in ('div', 'dividend'):
            continue
        friendly = DOLLAR_TAIL_PATTERN.sub('', detail.strip())
        lookup[ticker] = friendly  # most recent wins
    wb.close()
    return lookup


def infer_friendly_name(ticker, description):
    """For tickers not in history, infer 'Company Name (TICKER)' from description."""
    # description like "65266 TRELL HEALTH ORD  Del     .03 S Date 17/01/25"
    # strip leading qty number, drop the trading suffix
    name = re.sub(r'^\d+\s+', '', description or '')
    name = re.split(r'\s+Del\s+', name)[0]
    name = re.sub(r'\s+(ORD|COM SHS|NPV).*$', '', name)
    name = name.title().strip()
    if not name:
        name = ticker.title()
    return f"{name} ({ticker})"


# ============================================================================
# Build tracking-sheet row
# ============================================================================

def parse_div_qty(desc):
    m = re.match(r'Div\s+(\d+)\s', desc or '')
    return int(m.group(1)) if m else None


def parse_trade_qty(desc):
    m = re.match(r'^(\d+)\s', desc or '')
    return int(m.group(1)) if m else None


def build_tracking_row(src, ticker_lookup, flag_new_tickers):
    """Return a 26-element list ready to write into the tracking sheet."""
    cat = src['Category']
    nr = [None] * 26
    nr[0] = fmt_yyyymmdd(src['Date_parsed'])
    nr[10] = src['Running Balance_f']
    debit = src['Debit_f'] or 0
    credit = src['Credit_f'] or 0

    def get_friendly(ticker):
        if not ticker or ticker == 'n/a':
            return None
        if ticker in ticker_lookup:
            return ticker_lookup[ticker]
        # New ticker - infer and flag
        inferred = infer_friendly_name(ticker, src.get('Description', ''))
        flag_new_tickers.append((ticker, inferred, src.get('Description', '')))
        return inferred

    if cat == 'INTEREST':
        nr[1] = 'Account Interest'; nr[2] = 'Int'
        nr[5] = credit; nr[9] = credit; nr[18] = credit
    elif cat == 'DIVIDEND':
        qty = parse_div_qty(src['Description'])
        nr[1] = get_friendly(src['Symbol']) or src['Description']
        nr[2] = 'DIV'; nr[3] = qty
        if qty and qty > 0:
            nr[4] = round(credit / qty, 6)
        nr[5] = credit; nr[9] = credit; nr[16] = credit
    elif cat == 'WITHDRAWAL':
        nr[1] = 'BACS withdrawal'; nr[2] = 'Cash'
        nr[9] = debit; nr[13] = debit
    elif cat == 'DEPOSIT':
        nr[1] = 'A_Transfer from NatWest'; nr[2] = 'In'
        nr[9] = credit; nr[12] = credit
    elif cat == 'BUY':
        qty = src['Quantity_f'] or parse_trade_qty(src['Description'])
        price = src['Price_f']
        nr[1] = get_friendly(src['Symbol']) or (src.get('Description') or '')[:50]
        nr[2] = 'P'; nr[3] = qty; nr[4] = price
        if qty and price:
            sub = round(qty * price, 6); nr[5] = sub
            commn = round(debit - sub, 2)
            if abs(commn) > 0.005:
                nr[6] = commn
        nr[9] = debit; nr[20] = debit
        if qty and qty > 0:
            nr[11] = round(debit / qty, 6)
    elif cat == 'SELL':
        qty = src['Quantity_f'] or parse_trade_qty(src['Description'])
        price = src['Price_f']
        nr[1] = get_friendly(src['Symbol']) or (src.get('Description') or '')[:50]
        nr[2] = 'S'; nr[3] = qty; nr[4] = price
        # T = GROSS sales (qty x price), not the net Credit
        if qty and price:
            gross = round(qty * price, 6); nr[5] = gross
            nr[19] = gross  # T: Sales Total - GROSS
        else:
            nr[19] = credit  # fallback
        nr[9] = credit  # J = net proceeds (matches ii Credit)
    else:
        nr[1] = (src.get('Description') or '')[:50]
        nr[2] = 'OTHER'
    return nr


# ============================================================================
# XLSX writers (with null-pad workaround)
# ============================================================================

def safe_save_xlsx(wb, final_path):
    """Save XLSX to /tmp then cp to final to avoid Cowork null-pad bug."""
    tmp_path = Path(tempfile.gettempdir()) / Path(final_path).name
    wb.save(str(tmp_path))
    os.makedirs(os.path.dirname(final_path), exist_ok=True)
    shutil.copyfile(str(tmp_path), final_path)
    # Quick validity check: zip EOCD is 0x504b0506
    with open(final_path, 'rb') as f:
        f.seek(-22, os.SEEK_END)
        eocd_start = f.read(4)
    assert eocd_start[:4] == b'PK\x05\x06', f"XLSX EOCD missing in {final_path}"


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def write_extended_xlsx(existing_xlsx, output_path, header_text, opening_balance,
                        tracking_rows, summary_rows):
    """Load existing workbook, append YE block from max_row + 2."""
    wb = openpyxl.load_workbook(existing_xlsx)
    ws = wb["Sheet1"]
    
    # Find a generic prior-year tx row to use as style template.
    # Search backwards for the last row that has a value in col A (a tx row).
    tx_template = None
    for r in range(ws.max_row, 0, -1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and re.match(r'^\d{4}\.\d{2}\.\d{2}$', a):
            tx_template = r; break
    if tx_template is None:
        tx_template = ws.max_row  # fallback
    
    start_row = ws.max_row + 2  # blank separator then header
    
    # Header
    ws.cell(row=start_row, column=2, value=header_text)
    ws.cell(row=start_row, column=11, value=opening_balance)
    ws.cell(row=start_row, column=12, value=' < - From prior YE spreadsheet (after FX adj) - opening cash GBP')
    
    # Transaction rows
    for i, tr in enumerate(tracking_rows, 1):
        target = start_row + i
        for c, val in enumerate(tr, 1):
            if val is not None:
                ws.cell(row=target, column=c, value=val)
        # Copy template styles
        for c in range(1, 27):
            copy_style(ws.cell(row=tx_template, column=c), ws.cell(row=target, column=c))
    
    # Summary rows
    summary_start = start_row + len(tracking_rows) + 1
    for i, sr in enumerate(summary_rows):
        target = summary_start + i
        for c, val in enumerate(sr, 1):
            if val is not None:
                ws.cell(row=target, column=c, value=val)
        for c in range(1, 27):
            copy_style(ws.cell(row=tx_template, column=c), ws.cell(row=target, column=c))
    
    safe_save_xlsx(wb, output_path)
    return start_row, summary_start + len(summary_rows) - 1


def write_audit_xlsx(output_path, source_rows, tracking_rows, opening_balance,
                     totals, cash_position, stock_valuation, flag_new_tickers,
                     ye_date_str):
    """Build the 3-sheet audit XLSX."""
    wb = openpyxl.Workbook()
    
    # Styles
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    sub_font = Font(bold=True, size=10)
    sub_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    flag_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
    ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    final_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    
    # === Sheet 1: Transactions cross-check ===
    ws1 = wb.active; ws1.title = "Transactions"
    headers = ["ii Date", "ii Description", "ii Debit", "ii Credit", "ii Running Bal",
               "Category", "Sheet Date", "Sheet Detail", "Sheet Type",
               "Sheet J (Total)", "Sheet K (Bal)", "Bal check",
               "Q Divs", "S Int", "M Add", "N Withdr", "T Sales", "U Purch"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
    expected_bal = opening_balance
    for i, (src, tr) in enumerate(zip(source_rows, tracking_rows), start=2):
        if src['Credit_f']: expected_bal += src['Credit_f']
        if src['Debit_f']:  expected_bal -= src['Debit_f']
        expected_bal = round(expected_bal, 2)
        actual = src['Running Balance_f']
        bal_ok = abs(expected_bal - actual) < 0.005
        ws1.cell(row=i, column=1, value=src['Date_parsed'].strftime('%d/%m/%Y'))
        ws1.cell(row=i, column=2, value=src['Description'])
        ws1.cell(row=i, column=3, value=src['Debit_f'])
        ws1.cell(row=i, column=4, value=src['Credit_f'])
        ws1.cell(row=i, column=5, value=actual)
        ws1.cell(row=i, column=6, value=src['Category'])
        ws1.cell(row=i, column=7, value=tr[0])
        ws1.cell(row=i, column=8, value=tr[1])
        ws1.cell(row=i, column=9, value=tr[2])
        ws1.cell(row=i, column=10, value=tr[9])
        ws1.cell(row=i, column=11, value=tr[10])
        bc = ws1.cell(row=i, column=12, value="OK" if bal_ok else f"DRIFT {expected_bal:.2f}")
        bc.fill = ok_fill if bal_ok else flag_fill; bc.alignment = center
        ws1.cell(row=i, column=13, value=tr[16])
        ws1.cell(row=i, column=14, value=tr[18])
        ws1.cell(row=i, column=15, value=tr[12])
        ws1.cell(row=i, column=16, value=tr[13])
        ws1.cell(row=i, column=17, value=tr[19])
        ws1.cell(row=i, column=18, value=tr[20])
    # Totals row
    tot_r = len(source_rows) + 2
    ws1.cell(row=tot_r, column=2, value='YE totals').font = sub_font
    ws1.cell(row=tot_r, column=3, value=sum((r['Debit_f'] or 0) for r in source_rows))
    ws1.cell(row=tot_r, column=4, value=sum((r['Credit_f'] or 0) for r in source_rows))
    ws1.cell(row=tot_r, column=13, value=totals['Dividends'])
    ws1.cell(row=tot_r, column=14, value=totals['Interest'])
    ws1.cell(row=tot_r, column=15, value=totals['Added'])
    ws1.cell(row=tot_r, column=16, value=totals['Withdrawn'])
    ws1.cell(row=tot_r, column=17, value=totals['Sales'])
    ws1.cell(row=tot_r, column=18, value=totals['Purchases'])
    for c in range(1, 19):
        ws1.cell(row=tot_r, column=c).fill = sub_fill
        ws1.cell(row=tot_r, column=c).font = sub_font
    for c in [3, 4, 5, 10, 11, 13, 14, 15, 16, 17, 18]:
        for r in range(2, tot_r + 1):
            ws1.cell(row=r, column=c).number_format = '#,##0.00'
            ws1.cell(row=r, column=c).alignment = right
    widths = {1: 12, 2: 50, 3: 11, 4: 11, 5: 13, 6: 12, 7: 12, 8: 25, 9: 6,
              10: 12, 11: 13, 12: 14, 13: 10, 14: 10, 15: 10, 16: 10, 17: 10, 18: 10}
    for c, w in widths.items():
        ws1.column_dimensions[get_column_letter(c)].width = w
    ws1.freeze_panes = 'A2'
    
    # === Sheet 2: YE Summary ===
    ws2 = wb.create_sheet("YE Summary")
    ws2.cell(row=1, column=1, value=f"Interactive Investor (ii) - D.Box - YE {ye_date_str} Summary").font = Font(bold=True, size=14)
    
    row = 3
    ws2.cell(row=row, column=1, value=f"Cash position at {ye_date_str}").font = sub_font
    ws2.cell(row=row, column=1).fill = sub_fill; row += 1
    for label, val in [
        ("GBP cash (per ii running balance)", cash_position['gbp']),
        (f"EUR {cash_position['eur']} x {cash_position['eur_rate']}" if cash_position['eur_rate'] else "EUR (FLAG: rate/amount missing)",
         cash_position['eur_in_gbp']),
        (f"USD {cash_position['usd']} x {cash_position['usd_rate']}" if cash_position['usd_rate'] else "USD (FLAG: rate/amount missing)",
         cash_position['usd_in_gbp']),
        ("Total Cash (GBP equiv)", cash_position['total']),
    ]:
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=val)
        if "Total" in str(label):
            ws2.cell(row=row, column=1).font = sub_font; ws2.cell(row=row, column=2).font = sub_font
            ws2.cell(row=row, column=1).fill = final_fill; ws2.cell(row=row, column=2).fill = final_fill
        row += 1
    
    if stock_valuation:
        row += 1
        ws2.cell(row=row, column=1, value=f"Stock value at {ye_date_str}").font = sub_font
        ws2.cell(row=row, column=1).fill = sub_fill; row += 1
        for label, val in [
            ("Whole-account book cost", stock_valuation['book_cost']),
            ("Whole-account market value", stock_valuation['market_value']),
            ("Lower of cost or NRV (= YE stock value)", stock_valuation['lcnrv']),
        ]:
            ws2.cell(row=row, column=1, value=label)
            ws2.cell(row=row, column=2, value=val)
            if "Lower" in label:
                ws2.cell(row=row, column=1).font = sub_font; ws2.cell(row=row, column=2).font = sub_font
                ws2.cell(row=row, column=1).fill = final_fill; ws2.cell(row=row, column=2).fill = final_fill
            row += 1
    
    row += 1
    ws2.cell(row=row, column=1, value="YE P&L category totals (GBP)").font = sub_font
    ws2.cell(row=row, column=1).fill = sub_fill; row += 1
    for label, val in [
        ("Withdrawn (col N)", totals['Withdrawn']),
        ("Added (col M)", totals['Added']),
        ("Dividends (col Q)", totals['Dividends']),
        ("Interest (col S)", totals['Interest']),
        ("Sales Total (col T)", totals['Sales']),
        ("Purchase Total (col U)", totals['Purchases']),
    ]:
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=val); row += 1
    
    # Cash movement reconciliation
    row += 1
    ws2.cell(row=row, column=1, value="Cash movement reconciliation").font = sub_font
    ws2.cell(row=row, column=1).fill = sub_fill; row += 1
    computed_close = round(opening_balance + totals['Dividends'] + totals['Interest']
                           + totals['Added'] + totals['Sales']
                           - totals['Withdrawn'] - totals['Purchases'], 2)
    actual_close = source_rows[-1]['Running Balance_f'] if source_rows else 0
    diff = round(computed_close - actual_close, 2)
    for label, val in [
        ("Opening GBP cash", opening_balance),
        ("+ Dividends + Interest + Added + Sales", round(totals['Dividends'] + totals['Interest']
                                                          + totals['Added'] + totals['Sales'], 2)),
        ("- Withdrawn + Purchases", round(totals['Withdrawn'] + totals['Purchases'], 2)),
        ("Computed closing GBP cash", computed_close),
        ("ii reported closing GBP cash", actual_close),
        ("Reconciliation difference (should be 0)", diff),
    ]:
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=val)
        if "Reconciliation" in label:
            ws2.cell(row=row, column=1).font = sub_font; ws2.cell(row=row, column=2).font = sub_font
            ws2.cell(row=row, column=1).fill = final_fill if abs(diff) < 0.005 else flag_fill
            ws2.cell(row=row, column=2).fill = final_fill if abs(diff) < 0.005 else flag_fill
        row += 1
    
    # Journal lift-out checklist
    row += 1
    ws2.cell(row=row, column=1, value="Figures Mick to lift for the manual journals").font = sub_font
    ws2.cell(row=row, column=1).fill = sub_fill; row += 1
    journal_items = [
        f"1. Cash position to record in Xero at {ye_date_str}: GBP {cash_position['total']:.2f}"
        if cash_position.get('total') else
        f"1. Cash position to record in Xero at {ye_date_str}: FLAG cash position incomplete",
    ]
    if stock_valuation:
        journal_items.append(
            f"2. Stock value at {ye_date_str}: GBP {stock_valuation['lcnrv']:.2f} (whole-account LCNRV)"
        )
    journal_items += [
        f"3. Dividend income for the year: GBP {totals['Dividends']:.2f}",
        f"4. Interest income for the year: GBP {totals['Interest']:.2f}",
        f"5. Cash withdrawals from ii (sums col N): GBP {totals['Withdrawn']:.2f}",
        f"6. Purchases (sums col U): GBP {totals['Purchases']:.2f}",
        f"7. Sales (sums col T): GBP {totals['Sales']:.2f}",
    ]
    if flag_new_tickers:
        for t, inferred, desc in flag_new_tickers:
            journal_items.append(
                f"FLAG new ticker: {t} - inferred '{inferred}' from '{desc[:50]}' - confirm with Mick"
            )
    for j in journal_items:
        ws2.cell(row=row, column=1, value=j)
        ws2.cell(row=row, column=1).alignment = left
        row += 1
    
    for r in range(4, row + 1):
        cell = ws2.cell(row=r, column=2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    ws2.column_dimensions['A'].width = 70
    ws2.column_dimensions['B'].width = 18
    
    # === Sheet 3: Stock Valuation (if provided) ===
    if stock_valuation and stock_valuation.get('holdings'):
        ws3 = wb.create_sheet(f"Stock Valuation {ye_date_str}")
        ws3.cell(row=1, column=1, value=f"Stock Valuation as at {ye_date_str}").font = Font(bold=True, size=14)
        headers3 = ["Ticker", "Name", "Qty", "Market Value", "Book Cost", "Gain/Loss"]
        for c, h in enumerate(headers3, 1):
            cell = ws3.cell(row=3, column=c, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
        for i, h in enumerate(stock_valuation['holdings'], start=4):
            ws3.cell(row=i, column=1, value=h.get('ticker'))
            ws3.cell(row=i, column=2, value=h.get('name'))
            ws3.cell(row=i, column=3, value=h.get('qty'))
            ws3.cell(row=i, column=4, value=h.get('market_value'))
            ws3.cell(row=i, column=5, value=h.get('book_cost'))
            ws3.cell(row=i, column=6, value=h.get('gain_loss'))
        n = len(stock_valuation['holdings'])
        tot_row = n + 4
        ws3.cell(row=tot_row, column=2, value="Totals (GBP)").font = sub_font
        ws3.cell(row=tot_row, column=4, value=stock_valuation['market_value'])
        ws3.cell(row=tot_row, column=5, value=stock_valuation['book_cost'])
        for c in range(1, 7):
            ws3.cell(row=tot_row, column=c).fill = sub_fill; ws3.cell(row=tot_row, column=c).font = sub_font
        # LCNRV summary
        r = tot_row + 2
        ws3.cell(row=r, column=1, value="Whole-account LCNRV (= YE stock value)").font = sub_font
        ws3.cell(row=r, column=2, value=stock_valuation['lcnrv']); ws3.cell(row=r, column=2).font = sub_font
        ws3.cell(row=r, column=1).fill = final_fill; ws3.cell(row=r, column=2).fill = final_fill
        for c in [3, 4, 5, 6]:
            for rr in range(4, tot_row + 1):
                cell = ws3.cell(row=rr, column=c)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        ws3.column_dimensions['A'].width = 12; ws3.column_dimensions['B'].width = 36
        for c in [3, 4, 5, 6]: ws3.column_dimensions[get_column_letter(c)].width = 14
    
    safe_save_xlsx(wb, output_path)


# ============================================================================
# Portfolio PDF parser
# ============================================================================

def parse_portfolio_pdf(pdf_path):
    """Pull holdings from the ii portfolio valuation PDF. Returns dict
    with book_cost, market_value, lcnrv, holdings list. Best-effort -
    ii layout may change; returns None if parsing fails."""
    try:
        import pdfplumber
    except ImportError:
        print("pdfplumber not available; skipping portfolio PDF parse", file=sys.stderr)
        return None
    holdings = []
    total_mv = 0; total_bc = 0
    totals_line_mv = None; totals_line_bc = None
    money_pattern = re.compile(r"£[\-]?[\d,]+\.\d{2}")
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                for line in text.splitlines():
                    stripped = line.strip()
                    is_totals_line = stripped.startswith('GBP ') or stripped.startswith('Totals')
                    monies = money_pattern.findall(line)
                    
                    if is_totals_line:
                        # Totals row in ii PDF: 5 GBP-prefix monies expected.
                        # Order: idx 0=Day G/L, idx 1=MV, idx 2=MV dup, idx 3=BC, idx 4=G/L
                        if len(monies) >= 5:
                            try:
                                nums = [float(m.replace('£', '').replace(',', '')) for m in monies]
                                totals_line_mv = nums[1]
                                totals_line_bc = nums[3]
                            except ValueError:
                                pass
                        continue
                    
                    if len(monies) >= 4:
                        try:
                            parts = line.split()
                            ticker = parts[0] if parts and parts[0].isupper() and len(parts[0]) <= 6 else "(n/a)"
                            nums_clean = [float(m.replace('£', '').replace(',', '')) for m in monies]
                            # Per-row format: Day G/L, Day G/L%, MV, MV(dup), BC, GL, GL%
                            # We see 5 GBP-prefix monies: Day G/L, MV, MV dup, BC, GL
                            if len(nums_clean) >= 5:
                                mv = nums_clean[2]; bc = nums_clean[3]; gl = nums_clean[4]
                            else:
                                mv = nums_clean[1]; bc = nums_clean[2]; gl = mv - bc
                            holdings.append({
                                'ticker': ticker, 'name': ' '.join(parts[1:5]),
                                'market_value': mv, 'book_cost': bc, 'gain_loss': gl,
                            })
                            total_mv += mv; total_bc += bc
                        except (ValueError, IndexError):
                            continue
    except Exception as e:
        print(f"Portfolio PDF parse failed: {e}", file=sys.stderr)
        return None
    
    # Prefer the Totals line (authoritative); warn if per-row sums drift
    if totals_line_mv is not None and totals_line_bc is not None:
        if abs(total_mv - totals_line_mv) > 1.00 or abs(total_bc - totals_line_bc) > 1.00:
            print(f"WARN: per-row sums (MV {total_mv:.2f}, BC {total_bc:.2f}) drift from "
                  f"Totals line (MV {totals_line_mv:.2f}, BC {totals_line_bc:.2f}). Using Totals line.",
                  file=sys.stderr)
        total_mv = totals_line_mv
        total_bc = totals_line_bc
    
    if total_mv == 0 or total_bc == 0:
        return None
    return {
        'book_cost': round(total_bc, 2),
        'market_value': round(total_mv, 2),
        'lcnrv': round(min(total_bc, total_mv), 2),
        'holdings': holdings,
    }

# ============================================================================
# Workings note
# ============================================================================

def write_workings_note(output_path, ye_date_str, date_today, opening_balance,
                        source_rows, totals, cash_position, stock_valuation,
                        flag_new_tickers):
    actual_close = source_rows[-1]['Running Balance_f'] if source_rows else 0
    from collections import Counter
    cat_counts = Counter(r['Category'] for r in source_rows)
    cat_str = ', '.join(f"{n} {c.lower()}" for c, n in sorted(cat_counts.items()))
    
    flags_lines = []
    if not cash_position.get('eur_rate'):
        flags_lines.append("- EUR rate / amount not supplied - placeholder in spreadsheet")
    if not cash_position.get('usd_rate'):
        flags_lines.append("- USD rate / amount not supplied - placeholder in spreadsheet")
    if flag_new_tickers:
        for t, inferred, desc in flag_new_tickers:
            flags_lines.append(f"- NEW TICKER {t}: inferred '{inferred}' from '{desc[:60]}'. Confirm.")
    if not stock_valuation:
        flags_lines.append("- Stock valuation PDF not provided or unparseable - LCNRV not computed")
    
    content = f"""# Interactive Investor (ii) - YE {ye_date_str} workings note

Date: {date_today}
Generated by: ii-to-xero skill

## What ran

Parsed the ii transaction CSV ({len(source_rows)} rows), built a ticker
lookup from the existing tracking spreadsheet history, classified each
transaction, re-sorted chronologically (with reverse-CSV tie-break for
same-day rows), verified the running balance against ii, then wrote
the extended tracking XLSX, audit XLSX, and this note.

Classification: {cat_str}.

## Key figures

| Item | GBP |
|---|---|
| Opening cash | {opening_balance:.2f} |
| Closing cash (ii) | {actual_close:.2f} |
| Dividends | {totals['Dividends']:.2f} |
| Interest | {totals['Interest']:.2f} |
| Withdrawn | {totals['Withdrawn']:.2f} |
| Added | {totals['Added']:.2f} |
| Sales (gross) | {totals['Sales']:.2f} |
| Purchases | {totals['Purchases']:.2f} |

## Cash position at YE
"""
    if cash_position.get('total') is not None:
        content += f"""
| Component | GBP |
|---|---|
| GBP cash | {cash_position['gbp']:.2f} |
| EUR equivalent | {cash_position['eur_in_gbp']:.6f} |
| USD equivalent | {cash_position['usd_in_gbp']:.6f} |
| **Total Cash** | **{cash_position['total']:.6f}** |
"""
    else:
        content += "\n(Pending Mick's FX rates / USD cash to complete.)\n"
    
    if stock_valuation:
        content += f"""
## Stock value at YE (whole-account LCNRV basis)

- Total book cost: GBP {stock_valuation['book_cost']:.2f}
- Total market value: GBP {stock_valuation['market_value']:.2f}
- Lower of the two: GBP {stock_valuation['lcnrv']:.2f}

This is the YE stock value figure for Mick's manual stock journal.
"""
    
    if flags_lines:
        content += "\n## Flags for Mick\n\n" + "\n".join(flags_lines) + "\n"
    
    content += """
## Next steps

1. Mick to open the extended XLSX and eyeball the YE block.
2. Mick to lift figures from audit XLSX Sheet 2 for manual journals.
3. Stock value journal: compare YE stock value to prior-year stock
   value and journal the movement (Dr 350 / Cr 625 direction per
   prior-year handwritten template, or reverse if stock has gone up).
4. Commit the workspace from PowerShell with a session-descriptive
   commit message.
"""
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    # Markdown is safe to write directly (text - null pad bug only affects XLSX zip)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(content)


# ============================================================================
# Main
# ============================================================================

def main():
    ap = argparse.ArgumentParser(description="ii-to-xero: process Interactive Investor YE transactions")
    ap.add_argument('--csv', required=True, help='ii transactions CSV')
    ap.add_argument('--existing-xlsx', required=True, help='Prior-year tracking spreadsheet XLSX')
    ap.add_argument('--output-dir', required=True, help='Where to write the outputs')
    ap.add_argument('--opening-balance', required=True, type=float, help='YE opening GBP cash')
    ap.add_argument('--ye-date', required=True, help='YE date YYYY.MM.DD')
    ap.add_argument('--date-today', required=True, help='Today YYYY.MM.DD (for filenames)')
    ap.add_argument('--eur-rate', type=float, default=None)
    ap.add_argument('--usd-rate', type=float, default=None)
    ap.add_argument('--eur-cash', type=float, default=None)
    ap.add_argument('--usd-cash', type=float, default=None)
    ap.add_argument('--portfolio-pdf', default=None, help='Optional ii portfolio valuation PDF')
    args = ap.parse_args()
    
    # Parse
    print(f"Parsing {args.csv}...")
    source_rows = parse_csv(args.csv)
    print(f"  {len(source_rows)} transactions")
    
    # YE date helpers - YYYY.MM.DD for filenames, DD.MM.YYYY for display
    ye_date_compact = args.ye_date  # e.g. 2025.11.30
    ye_parts = args.ye_date.split('.')
    ye_date_uk = f"{ye_parts[2]}.{ye_parts[1]}.{ye_parts[0]}"  # e.g. 30.11.2025
    
    # Build ticker lookup from history
    print(f"Building ticker lookup from {args.existing_xlsx}...")
    ticker_lookup = build_ticker_lookup(args.existing_xlsx)
    print(f"  {len(ticker_lookup)} tickers in history")
    
    # Verify running balance
    print("Verifying running balance...")
    expected = args.opening_balance
    all_ok = True
    for r in source_rows:
        if r['Credit_f']: expected += r['Credit_f']
        if r['Debit_f']:  expected -= r['Debit_f']
        expected = round(expected, 2)
        if abs(expected - r['Running Balance_f']) >= 0.005:
            all_ok = False
            print(f"  DRIFT at {r['Date_parsed']}: expected {expected:.2f}, got {r['Running Balance_f']:.2f}",
                  file=sys.stderr)
    if not all_ok:
        print("Running balance reconciliation FAILED - aborting", file=sys.stderr)
        sys.exit(1)
    print(f"  CLEAN: ends at {expected:.2f}")
    
    # Build tracking rows
    flag_new_tickers = []
    tracking_rows = [build_tracking_row(r, ticker_lookup, flag_new_tickers) for r in source_rows]
    
    # Totals
    totals = {
        'Withdrawn': round(sum((tr[13] or 0) for tr in tracking_rows), 2),
        'Added':     round(sum((tr[12] or 0) for tr in tracking_rows), 2),
        'Dividends': round(sum((tr[16] or 0) for tr in tracking_rows), 2),
        'Interest':  round(sum((tr[18] or 0) for tr in tracking_rows), 2),
        'Sales':     round(sum((tr[19] or 0) for tr in tracking_rows), 2),
        'Purchases': round(sum((tr[20] or 0) for tr in tracking_rows), 2),
    }
    print(f"YE totals: {totals}")
    
    # Cash position
    closing_gbp = source_rows[-1]['Running Balance_f']
    eur_in_gbp = round(args.eur_cash * args.eur_rate, 6) if (args.eur_cash and args.eur_rate) else None
    usd_in_gbp = round(args.usd_cash * args.usd_rate, 6) if (args.usd_cash and args.usd_rate) else None
    total_cash = round(closing_gbp + (eur_in_gbp or 0) + (usd_in_gbp or 0), 6) \
                 if (eur_in_gbp is not None and usd_in_gbp is not None) else None
    cash_position = {
        'gbp': closing_gbp,
        'eur': args.eur_cash, 'eur_rate': args.eur_rate, 'eur_in_gbp': eur_in_gbp,
        'usd': args.usd_cash, 'usd_rate': args.usd_rate, 'usd_in_gbp': usd_in_gbp,
        'total': total_cash,
    }
    
    # Stock valuation
    stock_valuation = parse_portfolio_pdf(args.portfolio_pdf) if args.portfolio_pdf else None
    if stock_valuation:
        print(f"Stock: cost {stock_valuation['book_cost']}, market {stock_valuation['market_value']}, "
              f"LCNRV {stock_valuation['lcnrv']}")
    
    # Build the summary rows for the extended XLSX
    summary_rows = []
    # 1. Rounding adjustment (= 0 if reconciles clean)
    sr1 = [None] * 26
    sr1[1] = 'Rounding adjustment'
    sr1[9] = 0; sr1[10] = closing_gbp
    sr1[11] = f' <-- E. of Yr Balance (£), as at {ye_date_uk} - balance ties to ii'
    summary_rows.append(sr1)
    # 2. EUR
    sr2 = [None] * 26
    sr2[1] = 'Cash Held in Euros'
    if args.eur_cash is not None: sr2[5] = args.eur_cash
    sr2[6] = 'Euros'; sr2[7] = 'x '
    if args.eur_rate is not None: sr2[8] = args.eur_rate
    sr2[9] = '='
    sr2[10] = eur_in_gbp
    sr2[11] = ' <-- E. of Yr Balance (conversion from Euros)' if args.eur_rate else \
              ' <-- E. of Yr Balance (conversion from Euros) - FLAG: rate/amount pending'
    summary_rows.append(sr2)
    # 3. USD
    sr3 = [None] * 26
    sr3[1] = 'Cash Held in $US'
    if args.usd_cash is not None: sr3[5] = args.usd_cash
    sr3[6] = 'USD'; sr3[7] = 'x '
    if args.usd_rate is not None: sr3[8] = args.usd_rate
    sr3[9] = '='
    sr3[10] = usd_in_gbp
    sr3[11] = ' <-- E. of Yr Balance (conversion from $US)' if args.usd_rate else \
              ' <-- E. of Yr Balance (conversion from $US) - FLAG: rate/amount pending'
    summary_rows.append(sr3)
    # 4. Total Cash
    sr4 = [None] * 26
    sr4[1] = 'Total Cash (ii D.Box Account)'
    sr4[10] = total_cash
    sr4[11] = ' <-- E. of Yr Balance (Total expressed in £)' if total_cash is not None else \
              ' <-- E. of Yr Balance (Total expressed in £) - pending USD/EUR'
    summary_rows.append(sr4)
    # 5. YE Totals
    ye_year = args.ye_date.split('.')[0]
    sr5 = [None] * 26
    sr5[1] = f'YE 30.11.{ye_year}  :  Totals : £ '
    sr5[6] = 0; sr5[7] = 0
    sr5[12] = totals['Added']; sr5[13] = totals['Withdrawn']
    sr5[14] = 0; sr5[15] = 0
    sr5[16] = totals['Dividends']; sr5[17] = 0
    sr5[18] = totals['Interest']
    sr5[19] = totals['Sales']; sr5[20] = totals['Purchases']
    summary_rows.append(sr5)
    
    # Output filenames

    out_ext = Path(args.output_dir) / \
              f"{args.date_today} - D.Box_ii_Investment_SpSheet_YE_{ye_date_compact}_v.01.00_DRAFT.xlsx"
    out_audit = Path(args.output_dir) / \
                f"{args.date_today} - D.Box_ii_Audit_workings_YE_{ye_date_compact}_DRAFT.xlsx"
    out_note = Path(args.output_dir) / f"{args.date_today} - ii_Workings_Note.md"
    
    # Header text for the YE block in extended XLSX
    header_text = f'(Note: YE {ye_date_uk} - D.Box - Interactive Investor (ii) Account)'
    
    # Write
    print(f"Writing extended XLSX: {out_ext}")
    write_extended_xlsx(args.existing_xlsx, str(out_ext), header_text,
                        args.opening_balance, tracking_rows, summary_rows)
    
    print(f"Writing audit XLSX: {out_audit}")
    write_audit_xlsx(str(out_audit), source_rows, tracking_rows,
                     args.opening_balance, totals, cash_position,
                     stock_valuation, flag_new_tickers, ye_date_uk)
    
    print(f"Writing workings note: {out_note}")
    write_workings_note(str(out_note), ye_date_uk, args.date_today,
                        args.opening_balance, source_rows, totals,
                        cash_position, stock_valuation, flag_new_tickers)
    
    # Summary
    print("\n=== DONE ===")
    print(f"  Extended XLSX:  {out_ext}")
    print(f"  Audit XLSX:     {out_audit}")
    print(f"  Workings note:  {out_note}")
    if flag_new_tickers:
        print(f"\n  NEW TICKERS to confirm: {[t for t, _, _ in flag_new_tickers]}")
    return 0


if __name__ == '__main__':
    sys.exit(main())

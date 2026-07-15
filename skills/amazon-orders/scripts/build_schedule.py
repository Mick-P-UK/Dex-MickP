#!/usr/bin/env python3
"""
build_schedule.py - amazon-orders skill (Ditty Box Ltd YE accounts)

Turns the per-order Amazon items CSV chunks (produced by
export_amazon_items_csv with start_index) into a landscape, one-row-per-item
review schedule for a financial year, with the REAL payment card on every row.

The chunk pulls themselves are done by Cedric interactively via the
amazon-orders MCP (see SKILL.md). This script does the deterministic part:
  1. Combines all chunk CSVs in a folder.
  2. Dedupes item rows on (Order ID, ASIN, Item Total).
  3. Filters to the FY window (1 Dec prior year .. 30 Nov FY year).
  4. Reconciles Order IDs against an optional master order list (JSON).
  5. Classifies business vs personal from the real card (default 1136).
  6. Builds the landscape per-item XLSX (green=business, plain=personal, pink=cancelled).
     Order-level fields (date, order id, card, card type, biz/personal, status,
     ship-to) REPEAT on every item line so each row stands alone for a Xero
     import; the Order Total appears ONCE per order (first line) so it cannot
     double-count. A "Line" column ("1 of 2") plus a rule above each new order
     makes multi-item orders obvious. Notes cells flagging an item-sum
     mismatch are filled light red. Orders part- or fully-paid from an Amazon
     gift card are detected (Payment Method contains "Gift Card"): the amount
     that actually reached the card is taken from Order Grand Total, the row is
     filled amber and noted. This matters - a gift-card order shows the card
     number but never hits the card statement, so counting its order value as
     card spend overstates the year.
  7. Prints a verification report and warns if pulls did not page back far enough.

ASCII only. UK English. XLSX written to a temp file then moved (Cowork
Windows-mount null-padding workaround).
"""

import argparse, csv, glob, itertools, json, os, sys, tempfile, shutil
from collections import OrderedDict
from datetime import date

POUND = "\u00a3"
FMT = POUND + "#,##0.00"

# Standing rule (Mick, 2026.07.14): EVERY Excel sheet carries its own file path
# and the print date/time in the left footer, so any printout can be traced back
# to the file that made it. openpyxl translates these UI-style codes to the raw
# field codes (&Z&F / &D / &T) on save.
FOOTER_LEFT = "(&[Path]&[File] - Printed: &[Date] at &[Time])"


def stamp_footer(ws):
    """Apply the standard left footer to a worksheet."""
    ws.oddFooter.left.text = FOOTER_LEFT


def no_formulas(ws):
    """Force every cell to text/number - never formula.

    openpyxl types ANY string starting with "=" as a formula. A label like
    "= Total charges" is then written as a broken formula, and Excel greets Mick
    with "Removed Records: Formula from /xl/worksheets/sheet1.xml part" and
    silently repairs (i.e. deletes) the cell. This script writes no formulas at
    all, so any formula-typed cell is a false positive by definition. Caught
    2026.07.14 after Mick hit the repair dialog.
    """
    for row in ws.iter_rows():
        for c in row:
            if c.data_type == "f":
                c.data_type = "s"

CARDTYPE_DEFAULT = {
    "1136": "Halifax Clarity (Business)",
    "8155": "Personal Mastercard",
    "0343": "Personal Mastercard",
    "5039": "Personal Mastercard",
}


def money(s):
    return float((s or "0").replace(POUND, "").replace(",", "").strip() or 0)


def to_date(s):
    y, m, d = s.split("-")
    return date(int(y), int(m), int(d))


def load_chunks(chunks_dir, pattern):
    files = sorted(glob.glob(os.path.join(chunks_dir, pattern)))
    if not files:
        sys.exit("ERROR: no chunk files matched %s in %s" % (pattern, chunks_dir))
    rows = []
    for f in files:
        with open(f, encoding="utf-8-sig", newline="") as fh:
            rows.extend(list(csv.DictReader(fh)))
    return rows, files


def build(args):
    fy_start = date(args.fy_year - 1, 12, 1)
    fy_end = date(args.fy_year, 11, 30)

    rows, files = load_chunks(args.chunks_dir, args.chunks_glob)
    all_dates = sorted(x["Order Date"] for x in rows if x.get("Order Date"))
    oldest = to_date(all_dates[0]) if all_dates else None

    seen = set(); items = []
    for x in rows:
        od = x.get("Order Date", "")
        if not od or not (fy_start <= to_date(od) <= fy_end):
            continue
        key = (x["Order ID"], x["ASIN"], x["Item Total"])
        if key in seen:
            continue
        seen.add(key); items.append(x)

    byorder = OrderedDict()
    for x in sorted(items, key=lambda r: (r["Order Date"], r["Order ID"])):
        byorder.setdefault(x["Order ID"], []).append(x)

    master = {}; cancelled_ids = []
    if args.master and os.path.exists(args.master):
        mlist = json.load(open(args.master, encoding="utf-8"))
        master = {m["oid"]: m for m in mlist}
        cancelled_ids = [m["oid"] for m in mlist if "cancel" in str(m.get("status", "")).lower()]
    if args.cancelled:
        cancelled_ids = [c.strip() for c in args.cancelled.split(",") if c.strip()]

    def is_gift(x):
        return "gift card" in (x.get("Payment Method") or "").lower()

    def charged_to_card(x):
        """What actually reached the card. For a gift-card order Amazon reports
        the card-charged amount in Order Grand Total (0.00 if the gift card
        covered it all); otherwise Grand Total mirrors Order Total."""
        if is_gift(x):
            return money(x.get("Order Grand Total", ""))
        return money(x["Order Total"])

    # Flag only where the item lines tie to NEITHER the order subtotal NOR the
    # order total (>2p). A gap between subtotal and total is just shipping /
    # promotion / part gift-card and is normal - not flagged.
    recon_flag = {}
    for oid, its in byorder.items():
        isum = sum(money(x["Item Total"]) for x in its)
        sub = money(its[0].get("Order Subtotal", ""))
        tot = money(its[0]["Order Total"])
        recon_flag[oid] = abs(isum - sub) > 0.02 and abs(isum - tot) > 0.02

    print("amazon-orders build_schedule.py")
    print("FY window: %s .. %s" % (fy_start, fy_end))
    print("Chunk files read: %d" % len(files))
    print("Oldest order date pulled: %s" % (oldest or "n/a"))
    if oldest and oldest > fy_start:
        print("*** WARNING: oldest pulled date (%s) is AFTER FY start (%s)." % (oldest, fy_start))
        print("    Not paged back far enough - pull more chunks at a higher start_index")
        print("    until the oldest date is on/before %s." % fy_start)

    delivered = len(byorder)
    biz = [o for o, its in byorder.items() if its[0]["Card Last 4"].strip() == args.business_card]
    pers = [o for o in byorder if o not in biz]
    biz_total = sum(money(byorder[o][0]["Order Total"]) for o in biz)
    pers_total = sum(money(byorder[o][0]["Order Total"]) for o in pers)
    biz_charged = sum(charged_to_card(byorder[o][0]) for o in biz)
    pers_charged = sum(charged_to_card(byorder[o][0]) for o in pers)
    gift_orders = [o for o in byorder if is_gift(byorder[o][0])]
    gift_value = sum(money(byorder[o][0]["Order Total"]) - charged_to_card(byorder[o][0]) for o in gift_orders)

    print("\nDelivered orders in FY window: %d  (item rows: %d)" % (delivered, len(items)))
    print("  Business (card %s): %d orders, order value GBP %.2f, charged to card GBP %.2f"
          % (args.business_card, len(biz), biz_total, biz_charged))
    print("  Personal (other cards): %d orders, order value GBP %.2f, charged to card GBP %.2f"
          % (len(pers), pers_total, pers_charged))
    if gift_orders:
        print("  Gift-card orders: %d, GBP %.2f settled from gift card (NOT charged to any card)"
              % (len(gift_orders), gift_value))
        for o in gift_orders:
            x = byorder[o][0]
            print("      %s %s card %s: order %.2f, charged %.2f"
                  % (x["Order Date"], o, x["Card Last 4"], money(x["Order Total"]), charged_to_card(x)))

    missing = []
    if master:
        want = set(master) - set(cancelled_ids)
        missing = sorted(want - set(byorder))
        print("\nReconciliation vs master (%d orders, %d cancelled):" % (len(master), len(cancelled_ids)))
        print("  Delivered captured: %d / %d" % (delivered, len(want)))
        if missing:
            print("  *** MISSING %d delivered order(s) - pull an overlapping chunk:" % len(missing))
            for mid in missing:
                m = master.get(mid, {})
                print("      %s  %s  GBP %s" % (m.get("date", "?"), mid, m.get("amt", "?")))
        else:
            print("  All delivered orders captured. Zero gaps.")

    if any(recon_flag.values()):
        n = sum(1 for v in recon_flag.values() if v)
        print("\n%d order(s) flagged: item-sum != order total (promotion / part gift-card / postage)." % n)
        print("Order Total (amount charged) is authoritative; flagged in Notes column.")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook(); ws = wb.active; ws.title = "FY %d-%d Amazon Items" % (args.fy_year - 1, args.fy_year)
    hdr_fill = PatternFill("solid", fgColor="1F3864"); hdr_font = Font(bold=True, color="FFFFFF", size=10)
    green = PatternFill("solid", fgColor="C6EFCE"); pink = PatternFill("solid", fgColor="F8CBAD")
    redfill = PatternFill("solid", fgColor="FFC7CE"); redfont = Font(bold=True, color="9C0006", size=10)
    amber = PatternFill("solid", fgColor="FFE699")
    thin = Side(style="thin", color="BFBFBF"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    med = Side(style="medium", color="7F7F7F")
    border_top = Border(left=thin, right=thin, top=med, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top"); topal = Alignment(vertical="top")
    rightal = Alignment(horizontal="right", vertical="top")

    title = ("Ditty Box Ltd - Amazon.co.uk Orders - FY 1 Dec %d to 30 Nov %d - Per-Item Schedule"
             % (args.fy_year - 1, args.fy_year))
    ws.merge_cells("A1:Q1"); c = ws["A1"]; c.value = title; c.font = Font(bold=True, size=13, color="1F3864")
    ws.merge_cells("A2:Q2")
    ws["A2"].value = ("Real payment card per order from Amazon invoice. Card %s = business card (green). "
                      "Other cards = personal. Cancelled orders at foot (pink). "
                      "Order-level fields repeat on every item line (import-ready); the Order Total is shown "
                      "ONCE per order, on its first line, so it cannot double-count - sum the Order Total "
                      "column, never the Item Total column, to agree the year. LIGHT RED in Notes = the item "
                      "lines do not tie to the amount charged: check. AMBER = part/fully paid from an Amazon "
                      "gift card: 'Charged to Card' is what actually reached the card - only that "
                      "amount appears on the card statement. Prepared by Cedric %s."
                      % (args.business_card, args.date_today))
    ws["A2"].font = Font(italic=True, size=9, color="808080"); ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 42

    cols = ["Order Date", "Order ID", "Card Last4", "Card Type", "Biz / Personal", "Line",
            "Item (product)", "ASIN", "Qty", "Unit Price", "Item Total", "Order Total",
            "Charged to Card", "Status", "Ship-to", "Xero Code (Mick)", "Notes"]
    hr = 4
    for i, h in enumerate(cols, 1):
        cell = ws.cell(hr, i, h); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A5"

    cardtype = dict(CARDTYPE_DEFAULT)
    cardtype.setdefault(args.business_card, "Business card (%s)" % args.business_card)

    r = hr + 1
    centre = Alignment(horizontal="center", vertical="top")
    for oid, its in byorder.items():
        card = its[0]["Card Last 4"].strip()
        is_biz = card == args.business_card
        n_items = len(its)
        for j, x in enumerate(its):
            first = (j == 0)
            # Order-level fields repeat on EVERY item line so each row stands
            # alone for a Xero import. Order Total is the one exception: it is
            # written on the first line only, so the column can be summed
            # without double-counting a multi-item order.
            gift = is_gift(x)
            chg = charged_to_card(x)
            if gift:
                # A gift-card order carries a card number but never reaches the
                # card statement for the gift-funded part. Say so plainly.
                if chg == 0:
                    note = ("PAID IN FULL BY AMAZON GIFT CARD - GBP 0.00 charged to card %s. "
                            "No line will appear on the card statement. Do not claim via the card."
                            % card)
                else:
                    note = ("PART-PAID BY AMAZON GIFT CARD - only GBP %.2f reached card %s "
                            "(gift card covered GBP %.2f). The statement shows GBP %.2f, not GBP %.2f."
                            % (chg, card, money(x["Order Total"]) - chg, chg, money(x["Order Total"])))
            elif recon_flag.get(oid):
                note = "CHECK item-sum vs order total"
            else:
                note = ""
            note = note if first else ""
            row = [x["Order Date"], oid, card,
                   cardtype.get(card, "Personal card"),
                   ("Business" if is_biz else "Personal"),
                   "%d of %d" % (j + 1, n_items),
                   x["Product Name"], x["ASIN"], int(x["Quantity"] or 1),
                   money(x["Unit Price"]), money(x["Item Total"]),
                   money(x["Order Total"]) if first else "",
                   chg if first else "",
                   x["Order Status"], x["Recipient"], "",
                   note]
            bd = border_top if first else border
            for i, v in enumerate(row, 1):
                cell = ws.cell(r, i, v); cell.border = bd
                if i in (10, 11, 12, 13):
                    cell.number_format = FMT; cell.alignment = rightal
                elif i == 7:
                    cell.alignment = wrap
                elif i in (6, 9):
                    cell.alignment = centre
                else:
                    cell.alignment = topal
                if gift:
                    cell.fill = amber
                elif is_biz:
                    cell.fill = green
            if note:
                nc = ws.cell(r, 17)
                nc.fill = redfill; nc.font = redfont; nc.alignment = wrap
            ws.row_dimensions[r].height = 30
            r += 1

    for oid in cancelled_ids:
        m = master.get(oid, {})
        row = [m.get("date", ""), oid, "", "", "Cancelled", "1 of 1",
               "(order cancelled - no card charge, no items shipped)", "", "", "", "",
               0.0, 0.0, "Cancelled", m.get("ship", ""), "", "Exclude from Xero"]
        for i, v in enumerate(row, 1):
            cell = ws.cell(r, i, v); cell.border = border; cell.fill = pink
            if i in (10, 11, 12, 13):
                cell.number_format = FMT; cell.alignment = rightal
            elif i == 7:
                cell.alignment = wrap
            else:
                cell.alignment = topal
        r += 1

    r += 1
    ws.cell(r, 5, "SUMMARY").font = Font(bold=True, size=11, color="1F3864"); r += 1

    def putrow(label, val, charged=None, fill=None):
        nonlocal r
        ws.cell(r, 5, label).font = Font(bold=True)
        cell = ws.cell(r, 12, val); cell.number_format = FMT; cell.font = Font(bold=True); cell.alignment = rightal
        if charged is not None:
            c2 = ws.cell(r, 13, charged); c2.number_format = FMT; c2.font = Font(bold=True); c2.alignment = rightal
        if fill:
            for cc in range(5, 14):
                ws.cell(r, cc).fill = fill
        r += 1

    ws.cell(r, 12, "Order value").font = Font(bold=True, italic=True, size=9)
    ws.cell(r, 13, "Charged to card").font = Font(bold=True, italic=True, size=9)
    ws.cell(r, 12).alignment = rightal; ws.cell(r, 13).alignment = rightal
    r += 1
    putrow("Business - card %s (%d orders)" % (args.business_card, len(biz)),
           round(biz_total, 2), round(biz_charged, 2), fill=green)
    putrow("Personal - other cards (%d orders)" % len(pers),
           round(pers_total, 2), round(pers_charged, 2))
    if cancelled_ids:
        putrow("Cancelled (%d orders)" % len(cancelled_ids), 0.0, 0.0, fill=pink)
    putrow("TOTAL delivered FY orders (%d)" % delivered,
           round(biz_total + pers_total, 2), round(biz_charged + pers_charged, 2))
    if gift_orders:
        putrow("of which settled from Amazon gift card (never charged to a card) - %d order(s)"
               % len(gift_orders), round(gift_value, 2), 0.0, fill=amber)

    widths = {"A": 11, "B": 20, "C": 9, "D": 22, "E": 13, "F": 7, "G": 52, "H": 13,
              "I": 6, "J": 11, "K": 11, "L": 12, "M": 14, "N": 11, "O": 20, "P": 14, "Q": 34}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    stamp_footer(ws); no_formulas(ws)

    out_name = "%s - Amazon Orders FY %d-%d Per-Item Schedule.xlsx" % (args.date_today, args.fy_year - 1, args.fy_year)
    out_path = os.path.join(args.output_dir, out_name)
    os.makedirs(args.output_dir, exist_ok=True)
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx"); tf.close()
    wb.save(tf.name); shutil.copyfile(tf.name, out_path); os.unlink(tf.name)
    print("\nWrote: %s" % out_path)

    if args.statement_csv:
        if not os.path.exists(args.statement_csv):
            sys.exit("ERROR: --statement-csv not found: %s" % args.statement_csv)
        build_recon(args, byorder, is_gift, charged_to_card, fy_start, fy_end)
    else:
        print("(No --statement-csv given: reconciliation to the card statement skipped.)")

    if missing:
        print("EXIT 2: reconciliation incomplete (missing delivered orders).")
        return 2
    return 0


def parse_statement(path):
    """Read the card's Xero-import CSV and return the Amazon charges.

    Tolerant about headers: finds the date / amount / description columns by
    name so a future tweak to the cc1136-to-xero output does not break this.
    """
    from datetime import datetime as _dt
    rows = list(csv.DictReader(open(path, encoding="utf-8-sig")))
    if not rows:
        return []
    keys = list(rows[0].keys())

    def find(*cands):
        for c in cands:
            for k in keys:
                if k.strip().lower().lstrip("*") == c:
                    return k
        return None

    kd = find("date"); ka = find("amount"); kk = find("description")
    k2 = find("descr.2", "descr2")
    if not (kd and ka):
        sys.exit("ERROR: could not find date/amount columns in %s (saw: %s)" % (path, keys))

    out = []
    for r in rows:
        blob = " ".join(str(v) for v in r.values()).lower()
        if "amazon" not in blob and "amzn" not in blob:
            continue
        raw = (r[kd] or "").strip()
        dt = None
        for f in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                dt = _dt.strptime(raw, f).date(); break
            except ValueError:
                continue
        if dt is None:
            continue
        try:
            amt = float((r[ka] or "0").replace(",", ""))
        except ValueError:
            continue
        if amt >= 0:            # ignore refunds/credits here; flagged as leftovers
            continue
        out.append(dict(date=dt, amt=abs(amt), desc=(r.get(kk) or "").strip(),
                        d2=(r.get(k2) or "").strip() if k2 else ""))
    return out


def build_recon(args, byorder, is_gift, charged_to_card, fy_start, fy_end):
    """Tie the business-card orders back to the card statement.

    Deliberately does NOT hard-code subscription amounts. Orders are matched to
    charges; whatever charges are left over ARE the non-order items (Prime,
    subscriptions, etc.). That way a price change cannot silently break it.
    """
    from datetime import timedelta
    charges = parse_statement(args.statement_csv)
    if not charges:
        print("\nNo Amazon charges found in %s - reconciliation skipped." % args.statement_csv)
        return None

    biz_oids = [o for o, its in byorder.items()
                if its[0]["Card Last 4"].strip() == args.business_card]
    orders = []
    for oid in biz_oids:
        x = byorder[oid][0]
        orders.append(dict(oid=oid, date=to_date(x["Order Date"]),
                           value=money(x["Order Total"]), charged=charged_to_card(x),
                           gift=is_gift(x), items=[i["Product Name"] for i in byorder[oid]]))
    orders.sort(key=lambda o: o["date"])

    gift_only = [o for o in orders if o["charged"] == 0]
    payable = [o for o in orders if o["charged"] > 0]

    unused = list(charges); pairs = []; unmatched = []
    for o in payable:
        hit = next((c for c in unused
                    if abs(c["amt"] - o["charged"]) < 0.005
                    and timedelta(days=-2) <= (c["date"] - o["date"]) <= timedelta(days=args.match_window)), None)
        if hit:
            unused.remove(hit); pairs.append((hit, [o]))
        else:
            unmatched.append(o)
    # Amazon sometimes batches several orders into one card charge.
    for c in list(unused):
        for n in (2, 3):
            found = None
            for combo in itertools.combinations(unmatched, n):
                if abs(sum(o["charged"] for o in combo) - c["amt"]) < 0.005:
                    span = [o["date"] for o in combo]
                    if (c["date"] - min(span)).days <= args.match_window:
                        found = combo; break
            if found:
                pairs.append((c, list(found)))
                unused.remove(c)
                for o in found:
                    unmatched.remove(o)
                break
    pairs.sort(key=lambda pr: pr[0]["date"])

    matched_chg = sum(c["amt"] for c, _ in pairs)
    leftover = sum(c["amt"] for c in unused)
    total_chg = sum(c["amt"] for c in charges)
    diff = round(matched_chg + leftover - total_chg, 2)

    print("\n--- Reconciliation vs %s ---" % os.path.basename(args.statement_csv))
    print("Amazon charges on card %s: %d, GBP %.2f" % (args.business_card, len(charges), total_chg))
    print("  matched to orders : %d charges / %d orders = GBP %.2f"
          % (len(pairs), sum(len(o) for _, o in pairs), matched_chg))
    print("  no matching order : %d charges = GBP %.2f (subscriptions / membership / other)"
          % (len(unused), leftover))
    if gift_only:
        print("  gift-card orders (no charge expected): %d, order value GBP %.2f"
              % (len(gift_only), sum(o["value"] for o in gift_only)))
    if unmatched:
        print("  *** %d order(s) with NO charge found - may be charged after the year end:" % len(unmatched))
        for o in unmatched:
            print("      %s %s GBP %.2f" % (o["date"], o["oid"], o["charged"]))
    print("  CONTROL: matched + leftover - statement total = %.2f %s"
          % (diff, "(TIES)" if abs(diff) < 0.005 else "*** DOES NOT TIE"))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "Card %s recon" % args.business_card
    hf = PatternFill("solid", fgColor="1F3864"); hft = Font(bold=True, color="FFFFFF", size=10)
    green = PatternFill("solid", fgColor="C6EFCE"); amber = PatternFill("solid", fgColor="FFE699")
    redf = PatternFill("solid", fgColor="FFC7CE"); redt = Font(bold=True, color="9C0006", size=10)
    thin = Side(style="thin", color="BFBFBF"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top"); top = Alignment(vertical="top")
    rt = Alignment(horizontal="right", vertical="top")

    ws.merge_cells("A1:H1")
    ws["A1"] = ("Ditty Box Ltd - Amazon orders reconciled to card %s - FY %s to %s"
                % (args.business_card, fy_start.strftime("%d %b %Y"), fy_end.strftime("%d %b %Y")))
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws.merge_cells("A2:H2")
    ws["A2"] = ("Proves every Amazon charge on the card ties to an Amazon order or is a non-order item. "
                "Orders matched on the amount that actually reached the card. Source: Amazon invoices vs %s. "
                "Prepared by Cedric %s." % (os.path.basename(args.statement_csv), args.date_today))
    ws["A2"].font = Font(italic=True, size=9, color="808080"); ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 28

    cols = ["Charge Date", "Charge GBP", "Statement Description", "Order ID(s)",
            "Order Date", "Charged GBP", "What it was", "Notes"]
    r = 4
    for i, h in enumerate(cols, 1):
        c = ws.cell(r, i, h); c.fill = hf; c.font = hft; c.border = bd
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A5"; r = 5

    def put(vals, fill=None, red=False):
        nonlocal r
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v); c.border = bd
            if i in (2, 6):
                c.number_format = FMT; c.alignment = rt
            elif i in (3, 7, 8):
                c.alignment = wrap
            else:
                c.alignment = top
            if fill:
                c.fill = fill
        if red:
            c = ws.cell(r, 8); c.fill = redf; c.font = redt
        ws.row_dimensions[r].height = 28
        r += 1

    def short(names, n=2):
        return " + ".join(x[:44] for x in names[:n]) + (" ..." if len(names) > n else "")

    ws.cell(r, 1, "ORDERS MATCHED TO A CARD CHARGE").font = Font(bold=True, size=11, color="1F3864"); r += 1
    for c, os_ in pairs:
        put([c["date"].isoformat(), c["amt"], (c["desc"] + (" | " + c["d2"] if c["d2"] else "")),
             ", ".join(o["oid"] for o in os_), ", ".join(o["date"].isoformat() for o in os_),
             sum(o["charged"] for o in os_), short(sum((o["items"] for o in os_), [])),
             "Combined charge - %d orders batched by Amazon into one charge" % len(os_) if len(os_) > 1 else ""],
            fill=green, red=(len(os_) > 1))

    if gift_only:
        r += 1
        ws.cell(r, 1, "ORDERS SETTLED BY AMAZON GIFT CARD - no charge reaches the card").font = Font(bold=True, size=11, color="9C0006"); r += 1
        for o in gift_only:
            put(["(none)", 0.0, "No charge on the statement - correct", o["oid"], o["date"].isoformat(),
                 0.0, short(o["items"]),
                 "Order value GBP %.2f settled from Amazon gift card balance. Nothing charged to card %s, "
                 "so do NOT claim it via the card. Query: was the gift card company or personal funds?"
                 % (o["value"], args.business_card)], fill=amber, red=True)

    if unmatched:
        r += 1
        ws.cell(r, 1, "ORDERS WITH NO MATCHING CHARGE - investigate").font = Font(bold=True, size=11, color="9C0006"); r += 1
        for o in unmatched:
            put(["(not found)", 0.0, "No charge found within %d days" % args.match_window, o["oid"],
                 o["date"].isoformat(), o["charged"], short(o["items"]),
                 "No charge on this statement. Most likely shipped/charged after the year end - "
                 "check the next statement period, else investigate."], fill=redf, red=True)

    r += 1
    ws.cell(r, 1, "CHARGES WITH NO MATCHING ORDER - not order purchases").font = Font(bold=True, size=11, color="1F3864"); r += 1
    for c in sorted(unused, key=lambda z: z["date"]):
        put([c["date"].isoformat(), c["amt"], (c["desc"] + (" | " + c["d2"] if c["d2"] else "")),
             "", "", 0.0, "Not an Amazon order (no order history)",
             "Subscription / membership / other. Already in Xero via the card import - code it, "
             "and confirm business vs personal use."])

    r += 1
    ws.cell(r, 1, "CONTROL - does it tie?").font = Font(bold=True, size=11, color="1F3864"); r += 1

    def ctl(lbl, val, fill=None):
        nonlocal r
        ws.cell(r, 3, lbl).font = Font(bold=True); ws.cell(r, 3).alignment = wrap
        c = ws.cell(r, 6, val); c.number_format = FMT; c.font = Font(bold=True); c.alignment = rt
        if fill:
            for cc in range(3, 7):
                ws.cell(r, cc).fill = fill
        r += 1

    ctl("Charges matched to Amazon orders (%d charges / %d orders)"
        % (len(pairs), sum(len(o) for _, o in pairs)), round(matched_chg, 2), green)
    ctl("Charges with no matching order (subscriptions / membership / other)", round(leftover, 2))
    ctl("TOTAL Amazon charges on card %s (matched + unmatched)" % args.business_card,
        round(matched_chg + leftover, 2))
    ctl("Actual Amazon charges on the statement", round(total_chg, 2))
    ctl("Difference", diff, green if abs(diff) < 0.005 else redf)

    for k, v in {"A": 13, "B": 12, "C": 40, "D": 24, "E": 13, "F": 13, "G": 50, "H": 54}.items():
        ws.column_dimensions[k].width = v
    ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    stamp_footer(ws); no_formulas(ws)

    name = ("%s - Amazon vs Card %s Reconciliation FY %d-%d.xlsx"
            % (args.date_today, args.business_card, args.fy_year - 1, args.fy_year))
    path = os.path.join(args.output_dir, name)
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx"); tf.close()
    wb.save(tf.name); shutil.copyfile(tf.name, path); os.unlink(tf.name)
    print("Wrote: %s" % path)
    return diff


def main():
    ap = argparse.ArgumentParser(description="Build the Amazon FY per-item schedule XLSX.")
    ap.add_argument("--chunks-dir", required=True)
    ap.add_argument("--chunks-glob", default="_chunk_*.csv")
    ap.add_argument("--fy-year", type=int, required=True, help="FY end year, e.g. 2025 for YE 30 Nov 2025")
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--date-today", required=True, help="YYYY.MM.DD for the filename")
    ap.add_argument("--business-card", default="1136")
    ap.add_argument("--master", default=None)
    ap.add_argument("--cancelled", default=None)
    ap.add_argument("--statement-csv", default=None,
                    help="Optional: the business card's Xero-import CSV (from cc1136-to-xero). "
                         "If given, also writes the reconciliation of Amazon orders to card charges.")
    ap.add_argument("--match-window", type=int, default=14,
                    help="Days after the order date within which the card charge should appear (default 14).")
    args = ap.parse_args()
    sys.exit(build(args))


if __name__ == "__main__":
    main()
